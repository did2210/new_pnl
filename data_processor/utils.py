"""
Общие утилиты для обработки данных контрактов GFD.
"""
import re
import pandas as pd
from datetime import datetime, timedelta
from typing import Optional, Tuple
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment

from config import MONTH_NAMES_RU


# =============================================================================
# ПРЕОБРАЗОВАНИЕ ТИПОВ
# =============================================================================

def convert_to_string(value) -> str:
    """Преобразует любое значение в строку, обрабатывая кортежи и другие сложные типы."""
    if value is None:
        return ""
    if isinstance(value, tuple):
        if len(value) == 1:
            return str(value[0])
        return ", ".join([str(item) for item in value])
    return str(value)


def safe_to_float(value) -> float:
    """Безопасное преобразование значения в float."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return 0.0
    s = str(value).replace(chr(160), ' ').replace(' ', '').replace(',', '.').strip()
    if not s or s.lower() == 'nan' or s == '-':
        return 0.0
    try:
        return float(s)
    except Exception:
        return 0.0


def to_numeric_safe_with_null(series: pd.Series) -> pd.Series:
    """Безопасное приведение серии к числовому типу с обработкой NULL."""
    series_clean = series.astype(str).str.strip()
    series_clean = series_clean.replace(['NULL', 'null', 'Null', '', ' '], '0.0')
    series_clean = series_clean.str.replace(',', '.', regex=False)
    result = pd.to_numeric(series_clean, errors='coerce')
    return result.fillna(0.0)


# =============================================================================
# РАБОТА С ДАТАМИ
# =============================================================================

def parse_any_date(value) -> Tuple[str, Optional[datetime]]:
    """Парсит дату из любого формата."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return "", None

    if isinstance(value, datetime):
        return value.strftime('%d.%m.%Y'), value

    s = str(value).strip()
    if not s or s.lower() == 'nan':
        return "", None

    date_formats = ['%m/%d/%y', '%d.%m.%Y', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d']
    for fmt in date_formats:
        try:
            dt = datetime.strptime(s, fmt)
            return dt.strftime('%d.%m.%Y'), dt
        except ValueError:
            continue

    try:
        ts = pd.to_datetime(value, errors='coerce')
        if pd.notna(ts):
            dt = ts.to_pydatetime()
            return dt.strftime('%d.%m.%Y'), dt
    except Exception:
        pass

    return s, None


def get_russian_month_name(month_num: int) -> str:
    """Возвращает русское название месяца по номеру (1-12)."""
    if 1 <= month_num <= 12:
        return MONTH_NAMES_RU[month_num - 1]
    return ''


def month_name_to_num(month_name: str) -> Optional[int]:
    """Конвертирует русское название месяца в номер."""
    if not month_name:
        return None
    m = month_name.strip().lower()
    for idx, name in enumerate(MONTH_NAMES_RU):
        if name == m:
            return idx + 1
    return None


def get_contract_months(start_date: datetime, end_date: datetime) -> list:
    """Возвращает список (год, месяц) для всех месяцев контракта."""
    months = []
    current = start_date.replace(day=1)
    while current <= end_date:
        months.append((current.year, current.month))
        if current.month == 12:
            current = current.replace(year=current.year + 1, month=1)
        else:
            current = current.replace(month=current.month + 1)
    return months


# =============================================================================
# КАЛЕНДАРЬ ЭЦП
# =============================================================================

def build_ecp_calendar() -> dict:
    """
    Строит календарь недель ЭЦП на 2024-2027 годы.
    
    Правило:
    - W1 = дни 1-7 января
    - W2 = дни 8-14 января
    - Если в неделю попадает 1-е число месяца, вся неделя относится к этому месяцу
    
    Возвращает: {глобальный_номер_недели: (год, месяц, номер_недели_в_году)}
    """
    calendar = {}
    week_counter = 1

    for year in range(2024, 2028):
        for week_in_year in range(1, 53):
            start_day_of_year = (week_in_year - 1) * 7 + 1
            end_day_of_year = week_in_year * 7

            try:
                week_start = datetime(year, 1, 1) + timedelta(days=start_day_of_year - 1)
                year_end = datetime(year, 12, 31)
                week_end_candidate = datetime(year, 1, 1) + timedelta(days=end_day_of_year - 1)
                week_end = min(week_end_candidate, year_end)
            except Exception:
                break

            if week_start.year > year:
                break

            week_month = week_start.month
            week_year = week_start.year

            current_check_date = week_start
            while current_check_date <= week_end:
                if current_check_date.day == 1:
                    week_month = current_check_date.month
                    week_year = current_check_date.year
                    break
                current_check_date += timedelta(days=1)

            if week_year == year:
                calendar[week_counter] = (week_year, week_month, week_in_year)
                week_counter += 1
            else:
                break

    return calendar


def get_week_number_ecp(date: datetime) -> int:
    """Возвращает номер недели ЭЦП для даты."""
    day_of_year = date.timetuple().tm_yday
    week_num = ((day_of_year - 1) // 7) + 1
    return min(week_num, 52)


# =============================================================================
# КОПИРОВАНИЕ СТИЛЕЙ EXCEL
# =============================================================================

def safe_copy_style(source_cell, target_cell):
    """Безопасно копирует стили из одной ячейки в другую."""
    try:
        if source_cell.font:
            try:
                font_attrs = {}
                for attr in ['name', 'size', 'bold', 'italic', 'underline', 'color']:
                    val = getattr(source_cell.font, attr, None)
                    if val is not None:
                        font_attrs[attr] = val
                target_cell.font = Font(**font_attrs)
            except Exception:
                target_cell.font = Font()

        if source_cell.border:
            try:
                border_attrs = {}
                for side_name in ['left', 'right', 'top', 'bottom']:
                    side = getattr(source_cell.border, side_name, None)
                    if side and hasattr(side, 'style') and side.style:
                        border_attrs[side_name] = Side(style=side.style)
                target_cell.border = Border(**border_attrs)
            except Exception:
                target_cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )

        if source_cell.fill:
            try:
                if source_cell.fill.fill_type == "solid" and source_cell.fill.fgColor:
                    try:
                        rgb = source_cell.fill.fgColor.rgb
                        target_cell.fill = PatternFill(fill_type="solid", fgColor=rgb)
                    except Exception:
                        target_cell.fill = PatternFill(fill_type=None)
                else:
                    target_cell.fill = source_cell.fill
            except Exception:
                target_cell.fill = PatternFill(fill_type=None)

        if source_cell.alignment:
            try:
                alignment_attrs = {}
                for attr in ['horizontal', 'vertical', 'wrap_text', 'indent']:
                    val = getattr(source_cell.alignment, attr, None)
                    if val is not None:
                        alignment_attrs[attr] = val
                target_cell.alignment = Alignment(**alignment_attrs)
            except Exception:
                target_cell.alignment = Alignment(horizontal='left', vertical='center')

        if hasattr(source_cell, 'number_format') and source_cell.number_format:
            try:
                target_cell.number_format = source_cell.number_format
            except Exception:
                target_cell.number_format = 'General'
    except Exception:
        pass


# =============================================================================
# УТИЛИТЫ ДЛЯ EXCEL
# =============================================================================

def auto_fit_columns(ws, max_width=50):
    """Автоподбор ширины столбцов."""
    from openpyxl.utils import get_column_letter
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                try:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
                except Exception:
                    pass
        adjusted_width = min(max_length + 2, max_width)
        ws.column_dimensions[col_letter].width = max(adjusted_width, 8)


def add_thin_borders(ws, start_row, end_row, start_col, end_col):
    """Добавляет тонкие границы к диапазону ячеек."""
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            if not cell.border or (
                not cell.border.left.style and not cell.border.right.style and
                not cell.border.top.style and not cell.border.bottom.style
            ):
                cell.border = thin_border
