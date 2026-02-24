# -*- coding: utf-8 -*-
"""
Валидатор данных — проверки на каждом шаге пайплайна.

Категории проверок:
  1. ДАТЫ       — пустые, некорректные, start > end, pdate вне диапазона
  2. ПРОПУСКИ   — пустые viveska / sku / sap-code / filial
  3. ЗАТРАТЫ    — расхождение план/факт > 5×
  4. ОБЪЁМЫ     — отрицательные, аномально большие
  5. ЦЕНЫ       — нулевая цена при объёме > 0
  6. ДУБЛИКАТЫ  — повторы по ключевым полям
"""
import pandas as pd
from datetime import datetime
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .helpers import series_to_num

SEV_CRIT = "КРИТИЧЕСКАЯ"
SEV_WARN = "ПРЕДУПРЕЖДЕНИЕ"
SEV_INFO = "ИНФОРМАЦИЯ"

_COLORS = {SEV_CRIT: "CC0000", SEV_WARN: "CC8800", SEV_INFO: "0066CC"}


class Issue:
    __slots__ = ('sev', 'cat', 'file', 'sku', 'period', 'field',
                 'expected', 'actual', 'desc')

    def __init__(self, sev, cat, desc, **kw):
        self.sev = sev
        self.cat = cat
        self.desc = desc
        self.file = kw.get('file', '')
        self.sku = kw.get('sku', '')
        self.period = kw.get('period', '')
        self.field = kw.get('field', '')
        self.expected = kw.get('expected', '')
        self.actual = kw.get('actual', '')


def validate(df: pd.DataFrame) -> List[Issue]:
    """Запускает все проверки, возвращает список Issue."""
    if df is None or df.empty:
        return [Issue(SEV_CRIT, "Данные", "DataFrame пуст")]
    issues: List[Issue] = []
    _check_dates(df, issues)
    _check_missing(df, issues)
    _check_costs(df, issues)
    _check_volumes(df, issues)
    _check_prices(df, issues)
    _check_duplicates(df, issues)
    return issues


# ──────── проверки ────────

def _check_dates(df, out):
    for col in ('pdate', 'start_date', 'end_date'):
        if col not in df.columns:
            continue
        dt = pd.to_datetime(df[col], dayfirst=True, errors='coerce')
        bad = dt.isna() & df[col].notna() & (df[col].astype(str).str.strip() != '')
        if bad.any():
            for _, r in df[bad].head(5).iterrows():
                out.append(Issue(SEV_CRIT, "Даты",
                    f"Некорректная дата в '{col}': '{r.get(col)}'",
                    file=str(r.get('FileName', '')), sku=str(r.get('sku_type_sap', '')),
                    field=col, actual=str(r.get(col, ''))))
        empty = dt.isna() & (df[col].isna() | (df[col].astype(str).str.strip() == ''))
        if empty.any():
            out.append(Issue(SEV_WARN, "Даты",
                f"Пустая дата '{col}' у {empty.sum()} строк", field=col))

    if 'start_date' in df.columns and 'end_date' in df.columns:
        sd = pd.to_datetime(df['start_date'], dayfirst=True, errors='coerce')
        ed = pd.to_datetime(df['end_date'], dayfirst=True, errors='coerce')
        inv = (sd > ed) & sd.notna() & ed.notna()
        if inv.any():
            for _, r in df[inv].head(3).iterrows():
                out.append(Issue(SEV_CRIT, "Даты",
                    f"Дата начала ({r.get('start_date')}) > окончания ({r.get('end_date')})",
                    file=str(r.get('FileName', '')), field='start_date > end_date'))

    if all(c in df.columns for c in ('pdate', 'start_date', 'end_date')):
        p = pd.to_datetime(df['pdate'], dayfirst=True, errors='coerce')
        s = pd.to_datetime(df['start_date'], dayfirst=True, errors='coerce')
        e = pd.to_datetime(df['end_date'], dayfirst=True, errors='coerce')
        oor = ((p < s) | (p > e)) & p.notna() & s.notna() & e.notna()
        if oor.any():
            out.append(Issue(SEV_WARN, "Даты",
                f"pdate вне диапазона контракта у {oor.sum()} строк", field='pdate'))


def _check_missing(df, out):
    fields = {'viveska': 'Вывеска', 'sku_type_sap': 'SKU (SAP)',
              'sap-code': 'SAP-код', 'filial': 'Филиал'}
    for col, label in fields.items():
        if col not in df.columns:
            continue
        empty = df[col].isna() | (df[col].astype(str).str.strip().isin(['', 'nan', 'None']))
        if empty.any():
            sev = SEV_CRIT if col == 'sku_type_sap' else SEV_WARN
            files = df[empty]['FileName'].unique()[:5] if 'FileName' in df.columns else []
            out.append(Issue(sev, "Пропуски",
                f"Пустое поле '{label}' у {empty.sum()} строк",
                file=', '.join(str(f) for f in files), field=label))


def _check_costs(df, out):
    pairs = [
        ('Плановые затраты «Листинг/безусловные выплаты», руб',
         'Фактические затраты «Листинг/безусловные выплаты», руб', 'Листинг'),
        ('Плановые затраты «Скидка в цене», руб',
         'Фактические затраты «Скидка в цене», руб', 'Скидка в цене'),
        ('Плановые затраты «Ретро», руб',
         'Фактические затраты «Ретро», руб', 'Ретро'),
        ('Плановые затраты «Маркетинг», руб',
         'Фактические затраты «Маркетинг», руб', 'Маркетинг'),
        ('Плановые затраты «Промо-скидка», руб',
         'Фактические затраты «Промо-скидка», руб', 'Промо-скидка'),
    ]
    for pc, fc, label in pairs:
        if pc not in df.columns or fc not in df.columns:
            continue
        pv = series_to_num(df[pc])
        fv = series_to_num(df[fc])
        both = (pv > 0) & (fv > 0)
        if not both.any():
            continue
        ratio = fv[both] / pv[both]
        extreme = (ratio > 5) | (ratio < 0.2)
        if extreme.any():
            for _, r in df[both][extreme].head(5).iterrows():
                p, f = pv[r.name], fv[r.name]
                out.append(Issue(SEV_CRIT, "Затраты",
                    f"{label}: план={p:,.0f} факт={f:,.0f} (×{f/p:.1f})",
                    file=str(r.get('FileName', '')), sku=str(r.get('sku_type_sap', '')),
                    period=str(r.get('pdate', '')), field=label,
                    expected=f"{p:,.0f}", actual=f"{f:,.0f}"))


def _check_volumes(df, out):
    for col, label in [('volnew', 'Плановый объём'),
                        ('Плановые продажи, шт', 'Плановые продажи')]:
        if col not in df.columns:
            continue
        v = series_to_num(df[col])
        neg = v < 0
        if neg.any():
            out.append(Issue(SEV_CRIT, "Объёмы",
                f"Отрицательные значения '{label}' у {neg.sum()} строк", field=label))
        huge = v > 1_000_000
        if huge.any():
            out.append(Issue(SEV_INFO, "Объёмы",
                f"'{label}' > 1 000 000 у {huge.sum()} строк — проверьте", field=label))


def _check_prices(df, out):
    for pc, vc in [('price', 'volnew'), ('price_in', 'Плановые продажи, шт')]:
        if pc not in df.columns or vc not in df.columns:
            continue
        p = series_to_num(df[pc])
        v = series_to_num(df[vc])
        bad = (p == 0) & (v > 0)
        if bad.any():
            out.append(Issue(SEV_WARN, "Цены",
                f"Нулевая '{pc}' при объёме > 0 у {bad.sum()} строк", field=pc))


def _check_duplicates(df, out):
    keys = [c for c in ('FileName', 'viveska', 'sku_type_sap', 'pdate') if c in df.columns]
    if len(keys) < 2:
        return
    dups = df.duplicated(subset=keys, keep=False)
    if dups.any():
        groups = df[dups].groupby(keys).size()
        multi = groups[groups > 1]
        if len(multi):
            out.append(Issue(SEV_WARN, "Дубликаты",
                f"{len(multi)} групп дубликатов ({dups.sum()} строк)", field='ключевые поля'))


# ──────── отчёт в Excel ────────

def save_report(issues: List[Issue], path: str):
    wb = Workbook()
    # Сводка
    ws = wb.active
    ws.title = "Сводка"
    ws.merge_cells('A1:D1')
    ws['A1'] = "ОТЧЁТ ВАЛИДАЦИИ"
    ws['A1'].font = Font(bold=True, size=14, color="2F5496")
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A3'] = f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws['A3'].font = Font(italic=True)
    crit = sum(1 for i in issues if i.sev == SEV_CRIT)
    warn = sum(1 for i in issues if i.sev == SEV_WARN)
    info = sum(1 for i in issues if i.sev == SEV_INFO)
    for ri, (lbl, cnt, clr) in enumerate([
        (SEV_CRIT, crit, "CC0000"), (SEV_WARN, warn, "CC8800"), (SEV_INFO, info, "0066CC")
    ], 5):
        c1 = ws.cell(row=ri, column=1, value=lbl)
        c1.fill = PatternFill('solid', fgColor=clr)
        c1.font = Font(bold=True, color="FFFFFF")
        ws.cell(row=ri, column=2, value=cnt)
    ws.cell(row=9, column=1, value=f"Всего: {len(issues)}").font = Font(bold=True, size=12)
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12

    # Детали
    wd = wb.create_sheet("Детали")
    hdrs = ['Серьёзность', 'Категория', 'Файл', 'SKU', 'Период', 'Поле',
            'Ожидалось', 'Фактически', 'Описание']
    hfill = PatternFill('solid', fgColor="2F5496")
    hfont = Font(bold=True, color="FFFFFF", size=10)
    brd = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    for ci, h in enumerate(hdrs, 1):
        c = wd.cell(row=1, column=ci, value=h)
        c.fill = hfill
        c.font = hfont
        c.alignment = Alignment(horizontal='center', wrap_text=True)
        c.border = brd
    for ri, iss in enumerate(issues, 2):
        vals = [iss.sev, iss.cat, iss.file, iss.sku, iss.period,
                iss.field, iss.expected, iss.actual, iss.desc]
        for ci, v in enumerate(vals, 1):
            c = wd.cell(row=ri, column=ci, value=v)
            c.border = brd
            c.alignment = Alignment(vertical='center', wrap_text=True)
            if ci == 1:
                clr = _COLORS.get(iss.sev, "FFFFFF")
                c.fill = PatternFill('solid', fgColor=clr)
                c.font = Font(bold=True, color="FFFFFF")
    widths = [18, 22, 28, 22, 14, 22, 22, 22, 50]
    for ci, w in enumerate(widths, 1):
        wd.column_dimensions[get_column_letter(ci)].width = w
    wd.freeze_panes = 'A2'
    wb.save(path)
