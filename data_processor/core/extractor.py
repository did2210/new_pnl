# -*- coding: utf-8 -*-
"""
Шаг 1 — Извлечение таблиц из исходных «сырых» Excel-файлов.

Каждый исходник содержит несколько секций (GFD-запрос, условия контракта,
планирование продаж, планирование инвестиций, SAP-коды).
Мы ищем их по маркерам, вырезаем и складываем в единый «промежуточный» .xlsx
с отдельными листами.
"""
import os
import traceback
from datetime import datetime

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from .helpers import to_str, C, PLANNING_SHEETS

# ────────── внутренние утилиты ──────────

def _safe_copy_style(src, tgt):
    """Копируем стиль ячейки, игнорируя любые ошибки."""
    try:
        if src.font:
            tgt.font = Font(name=src.font.name, size=src.font.size,
                            bold=src.font.bold, italic=src.font.italic,
                            underline=src.font.underline, color=src.font.color)
    except Exception:
        pass
    try:
        if src.border:
            tgt.border = Border(left=src.border.left, right=src.border.right,
                                top=src.border.top, bottom=src.border.bottom)
    except Exception:
        pass
    try:
        if src.fill and src.fill.fill_type == 'solid' and src.fill.fgColor:
            tgt.fill = PatternFill(fill_type='solid', fgColor=src.fill.fgColor.rgb)
    except Exception:
        pass
    try:
        if src.alignment:
            tgt.alignment = Alignment(horizontal=src.alignment.horizontal,
                                      vertical=src.alignment.vertical,
                                      wrap_text=src.alignment.wrap_text)
    except Exception:
        pass
    try:
        if src.number_format:
            tgt.number_format = src.number_format
    except Exception:
        pass


def _find_text(wb, text):
    """Ищем текст по ВСЕМ листам. Возвращаем (row, col, sheet) или None."""
    for sn in wb.sheetnames:
        ws = wb[sn]
        for ri, row in enumerate(ws.iter_rows(values_only=False), 1):
            for ci, cell in enumerate(row, 1):
                if cell.value is not None:
                    if text in to_str(cell.value):
                        return ri, ci, ws
    return None


def _find_marker_row(ws, marker, start=1):
    """Номер строки, содержащей маркер (ищет по всем столбцам)."""
    for ri in range(start, ws.max_row + 1):
        for ci in range(1, min(ws.max_column + 1, 20)):
            v = ws.cell(row=ri, column=ci).value
            if v and isinstance(v, str) and marker in v:
                return ri
    return None


def _find_end(ws, start, stop_markers, max_empty=3):
    """
    Ищем конец секции: либо маркер начала другой секции, либо
    несколько подряд пустых строк.
    """
    empty_count = 0
    for ri in range(start, ws.max_row + 1):
        # маркер остановки?
        for ci in range(1, min(ws.max_column + 1, 20)):
            v = ws.cell(row=ri, column=ci).value
            if v and isinstance(v, str):
                for m in stop_markers:
                    if m.upper() in v.upper():
                        return ri - 1
        # пустая строка?
        empty = all(ws.cell(row=ri, column=c).value is None
                     for c in range(1, min(ws.max_column + 1, 30)))
        if empty:
            empty_count += 1
            if empty_count >= max_empty:
                return ri - max_empty
        else:
            empty_count = 0
    return ws.max_row


def _copy_block(ws, r1, r2, title, label):
    """Копируем диапазон строк из ws в новый Workbook. Возвращаем (wb, rows) или (None, 0)."""
    if r2 < r1:
        return None, 0
    wb_new = openpyxl.Workbook()
    ws_new = wb_new.active
    ws_new.title = title
    # заголовок
    ws_new.merge_cells('A1:Z1')
    h = ws_new['A1']
    h.value = label
    h.font = Font(bold=True, size=14)
    h.alignment = Alignment(horizontal='center')
    cur = 3
    rows_copied = 0
    for ri in range(r1, r2 + 1):
        has = False
        for ci in range(1, ws.max_column + 1):
            src = ws.cell(row=ri, column=ci)
            tgt = ws_new.cell(row=cur, column=ci)
            if src.value is not None:
                tgt.value = src.value
                has = True
            if src.has_style:
                _safe_copy_style(src, tgt)
        if has:
            cur += 1
            rows_copied += 1
    if rows_copied == 0:
        return None, 0
    # рамки
    thin = Border(left=Side('thin'), right=Side('thin'),
                  top=Side('thin'), bottom=Side('thin'))
    for r in range(3, cur):
        for c in range(1, ws.max_column + 1):
            cell = ws_new.cell(row=r, column=c)
            if not cell.border or not any(
                getattr(cell.border, s).style for s in ('left', 'right', 'top', 'bottom')
                if getattr(cell.border, s, None)
            ):
                cell.border = thin
    # ширина
    for ci in range(1, ws.max_column + 1):
        mx = 0
        for ri in range(3, cur):
            v = ws_new.cell(row=ri, column=ci).value
            if v:
                mx = max(mx, len(str(v)))
        ws_new.column_dimensions[get_column_letter(ci)].width = min(mx + 3, 50)
    return wb_new, rows_copied


def _planning_sheet(wb):
    """Находим лист с планированием."""
    for n in PLANNING_SHEETS:
        if n in wb.sheetnames:
            return wb[n], n
    return wb[wb.sheetnames[0]], wb.sheetnames[0]


# ════════════ ПУБЛИЧНЫЕ ФУНКЦИИ ════════════

class ExtractionResult:
    """Результат извлечения одного файла."""
    __slots__ = ('gfd_wb', 'gfd_rows', 'contract_wb', 'contract_rows',
                 'sales_wb', 'sales_rows', 'invest_wb', 'invest_rows',
                 'errors', 'warnings')
    def __init__(self):
        self.gfd_wb = self.contract_wb = self.sales_wb = self.invest_wb = None
        self.gfd_rows = self.contract_rows = self.sales_rows = self.invest_rows = 0
        self.errors = []
        self.warnings = []


def extract_tables(file_path: str) -> ExtractionResult:
    """
    Извлекает все секции из исходного файла.
    Возвращает ExtractionResult со всеми wb и счётчиками.
    """
    res = ExtractionResult()

    # ── GFD ──
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        hit = _find_text(wb, "Запрос на заключение контракта по напиткам GFD")
        if hit:
            ri, ci, ws = hit
            end = _find_end(ws, ri + 1,
                            ["Условия для нового контракта", "Условия контракта"])
            res.gfd_wb, res.gfd_rows = _copy_block(
                ws, ri + 1, end, "GFD Запрос",
                "ЗАПРОС НА ЗАКЛЮЧЕНИЕ КОНТРАКТА ПО НАПИТКАМ GFD")
        else:
            res.warnings.append("Секция 'GFD Запрос' не найдена")
    except Exception as e:
        res.errors.append(f"GFD: {e}")

    # ── Условия контракта ──
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        hit = _find_text(wb, "Условия для нового контракта")
        if hit:
            ri, ci, ws = hit
            # ищем "ВСЕГО:"
            end_r = None
            for r in range(ri + 1, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    v = ws.cell(row=r, column=c).value
                    if v and isinstance(v, str) and ("ВСЕГО:" in v or "ВСЕГО (" in v):
                        end_r = r - 1
                        break
                if end_r:
                    break
            if not end_r:
                end_r = _find_end(ws, ri + 1, [], max_empty=3)
            res.contract_wb, res.contract_rows = _copy_block(
                ws, ri + 1, end_r, "Условия контракта", "УСЛОВИЯ КОНТРАКТА")
        else:
            res.warnings.append("Секция 'Условия контракта' не найдена")
    except Exception as e:
        res.errors.append(f"Условия: {e}")

    # ── Планирование продаж ──
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws, sn = _planning_sheet(wb)
        start = _find_marker_row(ws, "Блок ПЛАНИРОВАНИЕ продаж")
        if start:
            end = _find_end(ws, start + 1,
                            ["Распределение инвестиций контракта", "ПЛАНИРОВАНИЕ ИНВЕСТИЦИЙ"])
            # пропускаем подзаголовок
            dr = start + 1
            for r in range(start, min(end + 1, ws.max_row + 1)):
                v = ws.cell(row=r, column=1).value
                if v and isinstance(v, str) and v.strip().upper() == "ПЛАНИРОВАНИЕ ПРОДАЖ":
                    dr = r + 1
                    break
            res.sales_wb, res.sales_rows = _copy_block(
                ws, dr, end, "Планирование продаж", "ПЛАНИРОВАНИЕ ПРОДАЖ")
        else:
            res.warnings.append("Секция 'Планирование продаж' не найдена")
    except Exception as e:
        res.errors.append(f"Продажи: {e}")

    # ── Планирование инвестиций ──
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws, sn = _planning_sheet(wb)
        start = _find_marker_row(ws, "Распределение инвестиций контракта, учитываемые в ЦМ")
        if not start:
            # fallback
            for r in range(1, ws.max_row + 1):
                v = ws.cell(row=r, column=1).value
                if v and isinstance(v, str) and v.strip().upper() == "ПЛАНИРОВАНИЕ ИНВЕСТИЦИЙ":
                    start = r
                    break
        if start:
            end = _find_end(ws, start + 1, ["Блок ПЛАНИРОВАНИЕ продаж"])
            dr = start + 1
            for r in range(start, min(end + 1, ws.max_row + 1)):
                v = ws.cell(row=r, column=1).value
                if v and isinstance(v, str) and v.strip().upper() == "ПЛАНИРОВАНИЕ ИНВЕСТИЦИЙ":
                    dr = r + 1
                    break
            res.invest_wb, res.invest_rows = _copy_block(
                ws, dr, end, "Планирование инвестиций", "ПЛАНИРОВАНИЕ ИНВЕСТИЦИЙ")
        else:
            res.warnings.append("Секция 'Планирование инвестиций' не найдена")
    except Exception as e:
        res.errors.append(f"Инвестиции: {e}")

    return res


def save_combined(res: ExtractionResult, src_path: str, out_path: str) -> bool:
    """Сохраняем все извлечённые листы + SAP-код в один файл."""
    try:
        merged = openpyxl.Workbook()
        merged.remove(merged.active)

        for wb, title in [
            (res.gfd_wb, "GFD Запрос"),
            (res.contract_wb, "Условия контракта"),
            (res.sales_wb, "Планирование продаж"),
            (res.invest_wb, "Планирование инвестиций"),
        ]:
            if wb:
                src_ws = wb.worksheets[0]
                new_ws = merged.create_sheet(title=title)
                for row in src_ws.iter_rows(values_only=False):
                    for cell in row:
                        nc = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                        _safe_copy_style(cell, nc)

        # SAP-код
        try:
            orig = openpyxl.load_workbook(src_path, data_only=False)
            if "SAP-код" in orig.sheetnames:
                sap = orig["SAP-код"]
                ns = merged.create_sheet(title="SAP-код")
                for row in sap.iter_rows(values_only=False):
                    for cell in row:
                        nc = ns.cell(row=cell.row, column=cell.column, value=cell.value)
                        _safe_copy_style(cell, nc)
        except Exception:
            res.warnings.append("Лист 'SAP-код' не скопирован")

        # сводка
        sm = merged.create_sheet(title="Сводка", index=0)
        sm.merge_cells('A1:D1')
        sm['A1'] = "ОБЪЕДИНЁННЫЙ ОТЧЁТ"
        sm['A1'].font = Font(bold=True, size=14)
        sm['A1'].alignment = Alignment(horizontal='center')
        for i, (lbl, cnt) in enumerate([
            ("GFD Запрос", res.gfd_rows),
            ("Условия контракта", res.contract_rows),
            ("Планирование продаж", res.sales_rows),
            ("Планирование инвестиций", res.invest_rows),
        ], 3):
            sm.cell(row=i, column=1, value=lbl).font = Font(bold=True)
            sm.cell(row=i, column=2, value=f"{cnt} строк")
        sm.column_dimensions['A'].width = 30
        sm.column_dimensions['B'].width = 15

        merged.save(out_path)
        return True
    except Exception as e:
        res.errors.append(f"Сохранение: {e}")
        return False
