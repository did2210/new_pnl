# -*- coding: utf-8 -*-
"""
Шаг 4 — Генерация итогового Excel P&L-отчёта.

Оптимизирован для больших объёмов (70к+ строк):
- pandas ExcelWriter с engine='openpyxl' для записи данных
- Форматирование ТОЛЬКО заголовков и ширины колонок (не каждой ячейки)
- Условное форматирование через openpyxl rules (а не цикл по ячейкам)
"""
import os
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

from .helpers import series_to_num, C, SKU_MAP, SKU_MAP_REV


def _fast_format(wb_path, row_count):
    """
    Быстрое форматирование: только заголовки + ширина + freeze.
    НЕ трогаем каждую ячейку данных — это убивает производительность на 70к+ строк.
    """
    print(f"  Форматирование...", end=" ", flush=True)
    wb = load_workbook(wb_path)

    hdr_fill = PatternFill('solid', fgColor="2F5496")
    hdr_font = Font(name='Calibri', size=10, bold=True, color="FFFFFF")
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    hdr_border = Border(left=Side('thin'), right=Side('thin'),
                        top=Side('thin'), bottom=Side('thin'))

    for ws in wb.worksheets:
        if ws.max_row <= 1:
            continue

        max_col = ws.max_column

        # 1. Форматируем ТОЛЬКО заголовок (строка 1)
        for ci in range(1, max_col + 1):
            c = ws.cell(row=1, column=ci)
            c.fill = hdr_fill
            c.font = hdr_font
            c.alignment = hdr_align
            c.border = hdr_border

        # 2. Ширина колонок (по первым 50 строкам, не по всем)
        sample_rows = min(ws.max_row, 50)
        for ci in range(1, max_col + 1):
            mx = 0
            for ri in range(1, sample_rows + 1):
                v = ws.cell(row=ri, column=ci).value
                if v is not None:
                    mx = max(mx, len(str(v)))
            ws.column_dimensions[get_column_letter(ci)].width = min(mx + 4, 35)

        # 3. Закрепляем шапку
        ws.freeze_panes = 'A2'

        # 4. Условное форматирование — красный для отрицательных в «Разница»
        #    (работает встроенными средствами Excel, не трогая каждую ячейку)
        red_font = Font(color="CC0000", bold=True)
        red_fill = PatternFill('solid', fgColor="FFF2F2")
        for ci in range(1, max_col + 1):
            header_val = str(ws.cell(row=1, column=ci).value or '')
            if 'разница' in header_val.lower():
                col_letter = get_column_letter(ci)
                cell_range = f"{col_letter}2:{col_letter}{ws.max_row}"
                ws.conditional_formatting.add(
                    cell_range,
                    CellIsRule(operator='lessThan', formula=['0'],
                               font=red_font, fill=red_fill)
                )

    wb.save(wb_path)
    print("OK")


def generate_report(merged_df: pd.DataFrame, out_path: str,
                    sales_path=None, costs_np_path=None, costs_ip_path=None,
                    cm_path=None, cogs_path=None) -> str:
    """
    Генерирует итоговый P&L-файл.
    """
    if merged_df.empty:
        print(C.warn("Нет данных для отчёта"))
        return None

    df = merged_df.copy()
    row_count = len(df)
    print(f"  Строк: {row_count:,}...", end=" ", flush=True)

    # ─── числовые ───
    num = ['price', 'price_in', 'listing', 'listing2', 'marketing', 'marketing2',
           'promo', 'promo2', 'retro', 'volnew', 'PromVol', 'dopmarketing']
    for c in num:
        if c in df.columns:
            df[c] = series_to_num(df[c])
    for c in ('pdate', 'start_date', 'end_date'):
        if c in df.columns:
            df[c] = pd.to_datetime(df[c], dayfirst=True, errors='coerce')

    # ─── группировка ───
    gk = [c for c in ('FileName', 'viveska', 'gr_sb', 'sku_type_sap',
                        'pdate', 'start_date', 'end_date') if c in df.columns]
    if 'volnew' in df.columns:
        plan = df.groupby(gk, as_index=False)['volnew'].sum()
        plan.rename(columns={'volnew': 'Плановые продажи, шт'}, inplace=True)
    else:
        plan = df[gk].drop_duplicates()
        plan['Плановые продажи, шт'] = 0
    dc = plan

    # ─── ЦМ ───
    if cm_path and os.path.exists(cm_path):
        try:
            cm = pd.read_excel(cm_path)
            cm['ЦМ'] = series_to_num(cm['ЦМ'])
            mk = [c for c in ('gr_sb', 'sku_type_sap', 'pdate') if c in cm.columns and c in dc.columns]
            if mk:
                dc = dc.merge(cm[mk + ['ЦМ']], on=mk, how='left')
                dc.rename(columns={'ЦМ': 'price_in'}, inplace=True)
        except Exception:
            pass
    if 'price_in' not in dc.columns and 'price_in' in df.columns:
        pi = df.groupby(gk, as_index=False)['price_in'].first()
        dc = dc.merge(pi, on=gk, how='left')
    if 'price_in' not in dc.columns:
        dc['price_in'] = 0.0
    dc['price_in'] = series_to_num(dc['price_in'])
    dc['Плановые продажи, руб'] = dc['Плановые продажи, шт'] * dc['price_in']

    # ─── инвестиционные колонки ───
    for c in ('listing2', 'listing', 'retro', 'dopmarketing', 'marketing', 'PromVol'):
        if c in df.columns:
            g = df.groupby(gk, as_index=False)[c].sum()
            dc = dc.merge(g, on=gk, how='left')
            dc[c] = series_to_num(dc[c])
        else:
            dc[c] = 0.0

    dc['Плановые затраты «Скидка в цене», руб'] = dc['Плановые продажи, руб'] * dc['listing2']
    dc['Плановые затраты «Листинг/безусловные выплаты», руб'] = dc['listing']
    dc['Плановые затраты «Ретро», руб'] = (dc['Плановые продажи, руб'] / 1.2) * dc['retro']
    dc['Плановые затраты «Маркетинг», руб'] = dc['Плановые продажи, руб'] * dc['dopmarketing'] + dc['marketing']
    dc['Плановые затраты «Промо-скидка», руб'] = dc['PromVol']

    # статус
    if 'pdate' in dc.columns and 'end_date' in dc.columns:
        dc['контракт'] = np.where(dc['pdate'] > dc['end_date'], 'завершенный', 'действующий')
    else:
        dc['контракт'] = 'действующий'

    # ─── Построение ecp_map: sap-code → viveska ───
    # (как load_ecp_map в оригинальном 4.py)
    ecp_map = pd.DataFrame()
    if 'sap-code' in df.columns and 'viveska' in df.columns:
        rows = []
        for _, r in df[['viveska', 'sap-code']].drop_duplicates().iterrows():
            codes = str(r['sap-code']).split(';')
            for code in codes:
                c = code.strip()
                if c and c.lower() not in ('nan', 'none', ''):
                    rows.append({'sap-code': c, 'viveska': r['viveska']})
        if rows:
            ecp_map = pd.DataFrame(rows)
            ecp_map['sap-code'] = series_to_num(ecp_map['sap-code'])
            ecp_map = ecp_map.drop_duplicates(subset=['sap-code', 'viveska'], keep='first')
        print(f"  ecp_map: {len(ecp_map)} записей (sap-code → viveska)")

    # ─── факт продаж ───
    # Sales.xlsx: zkcode, brand, sales_date, vol_2
    # zkcode → sap-code → viveska; brand → sku_type_sap; sales_date → pdate
    dc['Факт продажи, шт.'] = 0.0
    if sales_path and os.path.exists(sales_path):
        try:
            print(f"  Загрузка Sales...", end=" ", flush=True)
            sl = pd.read_excel(sales_path)
            sl['sales_date'] = pd.to_datetime(sl.get('sales_date'), errors='coerce')
            sl['zkcode'] = series_to_num(sl['zkcode']) if 'zkcode' in sl.columns else 0
            sl['vol_2'] = series_to_num(sl['vol_2']) if 'vol_2' in sl.columns else 0

            # Группируем
            if 'zkcode' in sl.columns and 'brand' in sl.columns:
                sa = sl.groupby(['zkcode', 'brand', 'sales_date'], as_index=False)['vol_2'].sum()

                # Привязываем viveska через ecp_map
                if not ecp_map.empty:
                    sa = sa.merge(ecp_map[['sap-code', 'viveska']],
                                  left_on='zkcode', right_on='sap-code', how='inner')
                    sa.drop(columns=['sap-code'], errors='ignore', inplace=True)

                # brand → sku_type_sap (оставляем как есть — названия совпадают)
                sa.rename(columns={'brand': 'sku_type_sap'}, inplace=True)
                sa['pdate'] = sa['sales_date']

                # Агрегируем по ключам
                sa_agg = sa.groupby(['viveska', 'sku_type_sap', 'pdate'], as_index=False)['vol_2'].sum()
                before = len(dc)
                dc = dc.merge(sa_agg, on=['viveska', 'sku_type_sap', 'pdate'], how='left')
                dc['Факт продажи, шт.'] = series_to_num(dc.get('vol_2', 0))
                dc.drop(columns=['vol_2'], errors='ignore', inplace=True)
                matched = (dc['Факт продажи, шт.'] > 0).sum()
                print(f"OK ({len(sa_agg)} записей, совпало {matched})")
            else:
                print("нет столбцов zkcode/brand")
        except Exception as e:
            print(f"ошибка: {e}")
    dc['Факт продажи, руб (от ЦМ)'] = dc['Факт продажи, шт.'] * dc['price_in']
    dc['Разница, шт'] = dc['Факт продажи, шт.'] - dc['Плановые продажи, шт']
    dc['Разница, руб'] = dc['Факт продажи, руб (от ЦМ)'] - dc['Плановые продажи, руб']

    # ─── факт затрат ───
    # затраты_вне_цены.xlsx: Номер заказчика → sap-code → viveska; Продукт → sku_type_sap
    # затраты_в_цене.xlsx: аналогично
    fact_cols = [
        'Фактические затраты «Листинг/безусловные выплаты», руб',
        'Фактические затраты «Ретро», руб',
        'Фактические затраты «Маркетинг», руб',
        'Фактические затраты «Промо-скидка», руб',
        'Фактические затраты «Скидка в цене», руб',
    ]
    if costs_np_path and os.path.exists(costs_np_path):
        try:
            print(f"  Загрузка затрат вне цены...", end=" ", flush=True)
            cnp = pd.read_excel(costs_np_path)
            cnp['Сумма'] = series_to_num(cnp['Сумма'])
            cnp['Номер заказчика'] = series_to_num(cnp['Номер заказчика'])
            cnp['pdate'] = pd.to_datetime(cnp.get('Месяц/год'), errors='coerce')
            if 'Продукт' in cnp.columns:
                cnp.rename(columns={'Продукт': 'sku_type_sap'}, inplace=True)
            # Привязываем viveska
            if not ecp_map.empty and 'Номер заказчика' in cnp.columns:
                cnp = cnp.merge(ecp_map[['sap-code', 'viveska']],
                                left_on='Номер заказчика', right_on='sap-code', how='left')
                cnp.drop(columns=['sap-code'], errors='ignore', inplace=True)
            # Фильтр: только «нет» в Фонды
            if 'Фонды' in cnp.columns:
                cnp = cnp[cnp['Фонды'].str.contains('нет', case=False, na=False)]
            mk = [c for c in ('pdate', 'viveska', 'sku_type_sap') if c in cnp.columns]
            if len(mk) == 3:
                for exp, cn in [('Листинг', fact_cols[0]), ('Маркетинг', fact_cols[2]),
                                ('Ретро', fact_cols[1])]:
                    f = cnp[cnp['Статья расходов'].str.contains(exp, case=False, na=False)]
                    if not f.empty:
                        g = f.groupby(mk, as_index=False)['Сумма'].sum()
                        dc = dc.merge(g, on=mk, how='left', suffixes=('', f'_{exp}'))
                        dc.rename(columns={'Сумма': cn}, inplace=True)
            print("OK")
        except Exception as e:
            print(f"ошибка: {e}")

    if costs_ip_path and os.path.exists(costs_ip_path):
        try:
            print(f"  Загрузка затрат в цене...", end=" ", flush=True)
            cip = pd.read_excel(costs_ip_path)
            cip['Сумма в валюте документа'] = series_to_num(cip['Сумма в валюте документа'])
            cip['Номер заказчика'] = series_to_num(cip['Номер заказчика'])
            cip['pdate'] = pd.to_datetime(cip.get('Месяц/год'), errors='coerce')
            if 'Продукт' in cip.columns:
                cip.rename(columns={'Продукт': 'sku_type_sap'}, inplace=True)
            # Привязываем viveska
            if not ecp_map.empty and 'Номер заказчика' in cip.columns:
                cip = cip.merge(ecp_map[['sap-code', 'viveska']],
                                left_on='Номер заказчика', right_on='sap-code', how='left')
                cip.drop(columns=['sap-code'], errors='ignore', inplace=True)
            mk = [c for c in ('pdate', 'viveska', 'sku_type_sap') if c in cip.columns]
            if len(mk) == 3 and 'примечание' in cip.columns:
                for pattern, cn in [('промо акция', fact_cols[3]), ('скидка в цене', fact_cols[4])]:
                    f = cip[cip['примечание'].str.contains(pattern, case=False, na=False)]
                    if not f.empty:
                        g = f.groupby(mk, as_index=False)['Сумма в валюте документа'].sum()
                        dc = dc.merge(g, on=mk, how='left', suffixes=('', f'_{pattern[:5]}'))
                        dc.rename(columns={'Сумма в валюте документа': cn}, inplace=True)
            print("OK")
        except Exception as e:
            print(f"ошибка: {e}")

    for c in fact_cols:
        if c not in dc.columns:
            dc[c] = 0.0
        else:
            dc[c] = series_to_num(dc[c])

    dc['план затраты'] = sum(dc.get(c, 0) for c in [
        'Плановые затраты «Листинг/безусловные выплаты», руб',
        'Плановые затраты «Скидка в цене», руб',
        'Плановые затраты «Ретро», руб',
        'Плановые затраты «Маркетинг», руб',
        'Плановые затраты «Промо-скидка», руб',
    ])
    dc['факт затраты'] = sum(dc.get(c, 0) for c in fact_cols)

    # ─── себестоимость ───
    if cogs_path and os.path.exists(cogs_path):
        try:
            cg = pd.read_excel(cogs_path)
            cg['cogs'] = series_to_num(cg['cogs'])
            mk = [c for c in ('sku_type_sap', 'pdate') if c in cg.columns and c in dc.columns]
            if mk:
                dc = dc.merge(cg[mk + ['cogs']], on=mk, how='left')
        except Exception:
            pass
    if 'cogs' not in dc.columns:
        dc['cogs'] = 0.0
    dc['cogs'] = series_to_num(dc['cogs'])
    dc['продажи по сс план'] = dc['Плановые продажи, шт'] * dc['cogs']
    dc['продажи по сс факт'] = dc['Факт продажи, шт.'] * dc['cogs']
    dc['доход план'] = dc['Плановые продажи, руб'] - dc['продажи по сс план'] - dc['план затраты']
    dc['доход факт'] = dc['Факт продажи, руб (от ЦМ)'] - dc['продажи по сс факт'] - dc['факт затраты']

    # дедуп
    dc.drop_duplicates(inplace=True)
    kk = [c for c in ('pdate', 'FileName', 'viveska', 'gr_sb', 'sku_type_sap',
                        'start_date', 'end_date') if c in dc.columns]
    if kk:
        dc.drop_duplicates(subset=kk, keep='first', inplace=True)

    # порядок
    order = [
        'pdate', 'FileName', 'viveska', 'gr_sb', 'sku_type_sap',
        'start_date', 'end_date', 'контракт',
        'Плановые продажи, шт', 'Факт продажи, шт.', 'Разница, шт',
        'Плановые продажи, руб', 'Факт продажи, руб (от ЦМ)', 'Разница, руб',
        'Плановые затраты «Листинг/безусловные выплаты», руб', fact_cols[0],
        'Плановые затраты «Скидка в цене», руб', fact_cols[4],
        'Плановые затраты «Ретро», руб', fact_cols[1],
        'Плановые затраты «Маркетинг», руб', fact_cols[2],
        'Плановые затраты «Промо-скидка», руб', fact_cols[3],
        'план затраты', 'факт затраты',
        'продажи по сс план', 'продажи по сс факт',
        'доход план', 'доход факт',
    ]
    order = [c for c in order if c in dc.columns]
    for c in dc.columns:
        if c not in order:
            order.append(c)
    dc = dc[order]

    final_rows = len(dc)
    print(f"расчёт OK ({final_rows:,} строк)")

    # ─── сохранение ───
    print(f"  Запись в Excel...", end=" ", flush=True)
    dc.to_excel(out_path, index=False, sheet_name='Результат', engine='openpyxl')
    print("OK")

    # ─── форматирование (быстрое — только заголовок) ───
    try:
        _fast_format(out_path, final_rows)
    except Exception as e:
        print(C.warn(f"Форматирование пропущено: {e}"))

    return out_path
