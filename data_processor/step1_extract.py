"""
Шаг 1: Извлечение таблиц из исходных Excel файлов.

Логика из оригинального кода "1.обработка исходников.py":
- Извлекает таблицу GFD запроса
- Извлекает условия контракта
- Извлекает планирование продаж
- Извлекает планирование инвестиций
- Копирует лист SAP-код
- Объединяет всё в один файл
"""
import os
import logging
import traceback
from datetime import datetime

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment

from utils import convert_to_string, safe_copy_style, auto_fit_columns, add_thin_borders
from config import TARGET_PLANNING_SHEETS

logger = logging.getLogger(__name__)


# =============================================================================
# ПОИСК ТЕКСТА В EXCEL
# =============================================================================

def find_text_in_excel(file_path, search_text, data_only=True):
    """
    Ищет указанный текст в Excel файле.
    Возвращает (номер строки, номер столбца, имя листа, ячейка, рабочая книга) или None.
    """
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=data_only)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row_idx, row in enumerate(sheet.iter_rows(values_only=False), 1):
                for col_idx, cell in enumerate(row, 1):
                    cell_value = cell.value
                    if cell_value is not None:
                        cell_value_str = convert_to_string(cell_value)
                        if isinstance(cell_value_str, str) and search_text in cell_value_str:
                            return (row_idx, col_idx, sheet_name, cell, workbook)
        return None
    except Exception as e:
        logger.error(f"Ошибка при поиске текста '{search_text}': {e}")
        return None


# =============================================================================
# КОПИРОВАНИЕ ДАННЫХ В НОВЫЙ ЛИСТ
# =============================================================================

def _copy_range_to_workbook(sheet, data_start_row, end_row, title, header_text, sheet_name):
    """Копирует диапазон строк из листа в новый workbook."""
    excel_wb = openpyxl.Workbook()
    excel_ws = excel_wb.active
    excel_ws.title = title

    # Заголовок
    excel_ws.merge_cells('A1:Z1')
    header_cell = excel_ws['A1']
    header_cell.value = header_text
    header_cell.font = Font(bold=True, size=16)
    header_cell.alignment = Alignment(horizontal='center')

    excel_ws.merge_cells('A2:Z2')
    info_text = (
        f"Дата извлечения: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | "
        f"Лист: {sheet_name} | "
        f"Диапазон: с {data_start_row} по {end_row}"
    )
    excel_ws['A2'] = info_text
    excel_ws['A2'].font = Font(italic=True)
    excel_ws['A2'].alignment = Alignment(horizontal='center')

    current_excel_row = 4
    rows_copied = 0

    for r in range(data_start_row, end_row + 1):
        has_data = False
        for c in range(1, sheet.max_column + 1):
            source_cell = sheet.cell(row=r, column=c)
            target_cell = excel_ws.cell(row=current_excel_row, column=c)
            if source_cell.value is not None:
                target_cell.value = source_cell.value
                has_data = True
            if source_cell.has_style:
                safe_copy_style(source_cell, target_cell)

        if has_data:
            current_excel_row += 1
            rows_copied += 1

    if rows_copied == 0:
        return None, 0

    # Добавляем рамки
    add_thin_borders(excel_ws, 4, current_excel_row - 1, 1, sheet.max_column)

    # Автоподбор ширины
    auto_fit_columns(excel_ws)

    # Подвал
    excel_ws.merge_cells(f'A{current_excel_row}:Z{current_excel_row}')
    excel_ws[f'A{current_excel_row}'] = f"Отчет сгенерирован автоматически {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    excel_ws[f'A{current_excel_row}'].font = Font(italic=True, size=10)
    excel_ws[f'A{current_excel_row}'].alignment = Alignment(horizontal='right')

    return excel_wb, rows_copied


# =============================================================================
# ИЗВЛЕЧЕНИЕ ТАБЛИЦ
# =============================================================================

def extract_gfd_request_table(file_path):
    """Извлекает таблицу 'Запрос на заключение контракта по напиткам GFD'."""
    try:
        result = find_text_in_excel(file_path, "Запрос на заключение контракта по напиткам GFD")
        if not result:
            logger.warning("Не найден текст 'Запрос на заключение контракта по напиткам GFD'")
            return None, 0

        row, col, sheet_name, cell, workbook = result
        sheet = workbook[sheet_name]
        logger.info(f"Найден заголовок GFD в {get_column_letter(col)}{row}")

        # Поиск конца таблицы
        end_row = None
        for r in range(row + 1, sheet.max_row + 1):
            for c in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=r, column=c)
                if cell.value and isinstance(cell.value, str):
                    if "Условия для нового контракта" in cell.value or "Условия контракта" in cell.value:
                        end_row = r - 1
                        break
            if end_row:
                break

        if not end_row:
            empty_row_count = 0
            for r in range(row + 1, sheet.max_row + 1):
                is_empty = all(sheet.cell(row=r, column=c).value is None for c in range(1, sheet.max_column + 1))
                if is_empty:
                    empty_row_count += 1
                    if empty_row_count >= 2:
                        end_row = r - 2
                        break
                else:
                    empty_row_count = 0
            if not end_row:
                end_row = sheet.max_row

        return _copy_range_to_workbook(
            sheet, row + 1, end_row,
            "GFD Запрос",
            "ЗАПРОС НА ЗАКЛЮЧЕНИЕ КОНТРАКТА ПО НАПИТКАМ GFD",
            sheet_name
        )
    except Exception as e:
        logger.error(f"Ошибка при извлечении GFD таблицы: {e}")
        traceback.print_exc()
        return None, 0


def extract_contract_conditions_table(file_path):
    """Извлекает таблицу условий контракта."""
    try:
        result = find_text_in_excel(file_path, "Условия для нового контракта", data_only=True)
        if not result:
            logger.warning("Не найден текст 'Условия для нового контракта'")
            return None, 0

        row, col, sheet_name, cell, workbook = result
        sheet = workbook[sheet_name]

        # Поиск ячейки "ВСЕГО:"
        total_cell = None
        for r in range(row, sheet.max_row + 1):
            for c in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=r, column=c)
                if cell.value is not None:
                    cell_value_str = convert_to_string(cell.value)
                    if isinstance(cell_value_str, str) and ("ВСЕГО:" in cell_value_str or "ВСЕГО (" in cell_value_str):
                        total_cell = (r, c, cell)
                        break
            if total_cell:
                break

        if not total_cell:
            logger.warning("Ячейка 'ВСЕГО:' не найдена")
            return None, 0

        return _copy_range_to_workbook(
            sheet, row + 1, total_cell[0] - 1,
            "Условия контракта",
            "УСЛОВИЯ КОНТРАКТА",
            sheet_name
        )
    except Exception as e:
        logger.error(f"Ошибка при извлечении условий контракта: {e}")
        traceback.print_exc()
        return None, 0


def extract_planning_sales_data(file_path):
    """Извлекает данные планирования продаж."""
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet, sheet_name = _find_planning_sheet(workbook)
        if not sheet:
            return None, 0

        start_marker = "Блок ПЛАНИРОВАНИЕ продаж"
        end_markers = [
            "Распределение инвестиций контракта, учитываемые в ЦМ, %",
            "ПЛАНИРОВАНИЕ ИНВЕСТИЦИЙ"
        ]

        start_row_idx = _find_marker_row(sheet, start_marker)
        if start_row_idx is None:
            logger.warning(f"Маркер '{start_marker}' не найден")
            return None, 0

        end_row_idx = _find_end_marker_row(sheet, start_row_idx + 1, end_markers)
        if end_row_idx is None:
            end_row_idx = sheet.max_row

        # Ищем подзаголовок
        data_start_row = start_row_idx + 1
        for r in range(start_row_idx, min(end_row_idx + 1, sheet.max_row + 1)):
            first_cell = sheet.cell(row=r, column=1)
            if first_cell.value and isinstance(first_cell.value, str) and \
               first_cell.value.strip().upper() == "ПЛАНИРОВАНИЕ ПРОДАЖ":
                data_start_row = r + 1
                break

        return _copy_range_to_workbook(
            sheet, data_start_row, end_row_idx,
            "Планирование продаж",
            "ПЛАНИРОВАНИЕ ПРОДАЖ",
            sheet_name
        )
    except Exception as e:
        logger.error(f"Ошибка при извлечении данных планирования продаж: {e}")
        traceback.print_exc()
        return None, 0


def extract_investment_planning_data(file_path):
    """Извлекает данные планирования инвестиций."""
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet, sheet_name = _find_planning_sheet(workbook)
        if not sheet:
            return None, 0

        start_marker = "Распределение инвестиций контракта, учитываемые в ЦМ, %"
        end_markers = ["Блок ПЛАНИРОВАНИЕ продаж"]

        start_row_idx = _find_marker_row(sheet, start_marker)

        if start_row_idx is None:
            # Альтернативный поиск
            for row_idx in range(1, sheet.max_row + 1):
                first_cell = sheet.cell(row=row_idx, column=1)
                if first_cell.value and isinstance(first_cell.value, str) and \
                   first_cell.value.strip().upper() == "ПЛАНИРОВАНИЕ ИНВЕСТИЦИЙ":
                    start_row_idx = row_idx
                    break

        if start_row_idx is None:
            logger.warning("Маркер начала инвестиций не найден")
            return None, 0

        end_row_idx = _find_end_marker_row(sheet, start_row_idx + 1, end_markers)
        if end_row_idx is None:
            end_row_idx = sheet.max_row

        data_start_row = start_row_idx + 1
        for r in range(start_row_idx, min(end_row_idx + 1, sheet.max_row + 1)):
            first_cell = sheet.cell(row=r, column=1)
            if first_cell.value and isinstance(first_cell.value, str) and \
               first_cell.value.strip().upper() == "ПЛАНИРОВАНИЕ ИНВЕСТИЦИЙ":
                data_start_row = r + 1
                break

        return _copy_range_to_workbook(
            sheet, data_start_row, end_row_idx,
            "Планирование инвестиций",
            "ПЛАНИРОВАНИЕ ИНВЕСТИЦИЙ",
            sheet_name
        )
    except Exception as e:
        logger.error(f"Ошибка при извлечении данных инвестиций: {e}")
        traceback.print_exc()
        return None, 0


# =============================================================================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# =============================================================================

def _find_planning_sheet(workbook):
    """Находит лист с данными планирования."""
    for name in TARGET_PLANNING_SHEETS:
        if name in workbook.sheetnames:
            return workbook[name], name
    sheet_name = workbook.sheetnames[0]
    return workbook[sheet_name], sheet_name


def _find_marker_row(sheet, marker_text):
    """Ищет строку с указанным маркером."""
    for row_idx, row in enumerate(sheet.iter_rows(values_only=False), 1):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and marker_text in cell.value:
                return row_idx
    return None


def _find_end_marker_row(sheet, start_from, markers):
    """Ищет строку с маркером конца секции."""
    for row_idx in range(start_from, sheet.max_row + 1):
        first_cell = sheet.cell(row=row_idx, column=1)
        if first_cell.value and isinstance(first_cell.value, str):
            cell_val_upper = first_cell.value.strip().upper()
            for marker in markers:
                if marker.strip().upper() in cell_val_upper:
                    return row_idx - 1
    return None


# =============================================================================
# ОБЪЕДИНЕНИЕ И ОСНОВНАЯ ФУНКЦИЯ
# =============================================================================

def merge_extracted_tables(gfd_wb, contract_wb, planning_wb, investment_wb, source_file_path, output_path):
    """Объединяет все извлечённые таблицы в один Excel файл."""
    try:
        merged_wb = openpyxl.Workbook()
        default_sheet = merged_wb.active
        merged_wb.remove(default_sheet)

        tables = [
            (gfd_wb, "GFD Запрос"),
            (contract_wb, "Условия контракта"),
            (planning_wb, "Планирование продаж"),
            (investment_wb, "Планирование инвестиций"),
        ]

        for wb, title in tables:
            if wb:
                src_ws = list(wb.worksheets)[0]
                new_sheet = merged_wb.create_sheet(title=title)
                for row in src_ws.iter_rows(values_only=False):
                    for cell in row:
                        new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                        safe_copy_style(cell, new_cell)

        # Копируем лист SAP-код
        try:
            source_wb = openpyxl.load_workbook(source_file_path, data_only=False)
            if "SAP-код" in source_wb.sheetnames:
                sap_sheet = source_wb["SAP-код"]
                new_sap_sheet = merged_wb.create_sheet(title="SAP-код")
                for row in sap_sheet.iter_rows(values_only=False):
                    for cell in row:
                        new_cell = new_sap_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                        safe_copy_style(cell, new_cell)
                logger.info("Лист 'SAP-код' скопирован")
        except Exception as e:
            logger.warning(f"Не удалось скопировать лист 'SAP-код': {e}")

        # Сводный лист
        summary_sheet = merged_wb.create_sheet(title="Сводка", index=0)
        summary_sheet.merge_cells('A1:D1')
        summary_sheet['A1'] = "ОБЪЕДИНЕННЫЙ ОТЧЕТ ПО КОНТРАКТАМ"
        summary_sheet['A1'].font = Font(bold=True, size=16)
        summary_sheet['A1'].alignment = Alignment(horizontal='center')

        info_rows = [
            ("A3", "Таблица 1:", "B3", "Запрос на заключение контракта по напиткам GFD"),
            ("A4", "Таблица 2:", "B4", "Условия контракта"),
            ("A5", "Таблица 3:", "B5", "Планирование продаж"),
            ("A6", "Таблица 4:", "B6", "Планирование инвестиций"),
            ("A7", "Лист 5:", "B7", "SAP-код"),
        ]
        for label_cell, label_text, value_cell, value_text in info_rows:
            summary_sheet[label_cell] = label_text
            summary_sheet[label_cell].font = Font(bold=True)
            summary_sheet[value_cell] = value_text

        summary_sheet.column_dimensions['A'].width = 15
        summary_sheet.column_dimensions['B'].width = 50

        merged_wb.save(output_path)
        logger.info(f"Объединённый файл сохранён: {output_path}")
        return True
    except Exception as e:
        logger.error(f"Ошибка при объединении: {e}")
        traceback.print_exc()
        return False


def process_raw_file(file_path, output_dir):
    """
    Обрабатывает один исходный Excel файл: извлекает все таблицы и объединяет.
    Возвращает путь к выходному файлу или None при ошибке.
    """
    logger.info(f"Обработка исходного файла: {os.path.basename(file_path)}")

    gfd_wb, gfd_rows = extract_gfd_request_table(file_path)
    logger.info(f"  GFD: {gfd_rows} строк" if gfd_wb else "  GFD: не найдено")

    contract_wb, contract_rows = extract_contract_conditions_table(file_path)
    logger.info(f"  Условия: {contract_rows} строк" if contract_wb else "  Условия: не найдено")

    planning_wb, planning_rows = extract_planning_sales_data(file_path)
    logger.info(f"  Планирование продаж: {planning_rows} строк" if planning_wb else "  Планирование продаж: не найдено")

    investment_wb, investment_rows = extract_investment_planning_data(file_path)
    logger.info(f"  Инвестиции: {investment_rows} строк" if investment_wb else "  Инвестиции: не найдено")

    base_name = os.path.basename(file_path)
    output_path = os.path.join(output_dir, base_name)

    success = merge_extracted_tables(
        gfd_wb, contract_wb, planning_wb, investment_wb, file_path, output_path
    )

    return output_path if success else None
