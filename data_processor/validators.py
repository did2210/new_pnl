"""
Валидаторы данных: проверки на расхождения, ошибки в датах, пропущенные данные.

Выполняет следующие проверки:
1. Расхождения в затратах: плановые затраты из кода (контракта) != затраты из файла
2. Проблемы с датами: пустые, некорректные, даты начала > даты окончания
3. Пропущенные данные: SKU без цены, без объёмов, без SAP-кодов
4. Аномалии: слишком большие/малые значения, нулевые цены при ненулевых объёмах
"""
import logging
from datetime import datetime
from typing import Optional

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from utils import to_numeric_safe_with_null

logger = logging.getLogger(__name__)

# Уровни серьёзности
SEVERITY_CRITICAL = "КРИТИЧЕСКАЯ"
SEVERITY_WARNING = "ПРЕДУПРЕЖДЕНИЕ"
SEVERITY_INFO = "ИНФОРМАЦИЯ"

# Цвета для серьёзности
SEVERITY_COLORS = {
    SEVERITY_CRITICAL: "FF4444",
    SEVERITY_WARNING: "FFAA00",
    SEVERITY_INFO: "4488FF",
}


class ValidationIssue:
    """Представляет одну найденную проблему."""

    def __init__(self, severity: str, category: str, file_name: str, sku: str,
                 period: str, field: str, expected: str, actual: str, description: str):
        self.severity = severity
        self.category = category
        self.file_name = file_name
        self.sku = sku
        self.period = period
        self.field = field
        self.expected = expected
        self.actual = actual
        self.description = description

    def to_dict(self):
        return {
            'Серьёзность': self.severity,
            'Категория': self.category,
            'Файл': self.file_name,
            'SKU': self.sku,
            'Период': self.period,
            'Поле': self.field,
            'Ожидалось': self.expected,
            'Фактически': self.actual,
            'Описание': self.description,
        }


class DataValidator:
    """Класс для выполнения всех проверок данных."""

    def __init__(self):
        self.issues: list[ValidationIssue] = []

    def validate_all(self, df: pd.DataFrame) -> list[ValidationIssue]:
        """Выполняет все проверки и возвращает список проблем."""
        self.issues = []

        if df is None or df.empty:
            self.issues.append(ValidationIssue(
                severity=SEVERITY_CRITICAL,
                category="Данные",
                file_name="",
                sku="",
                period="",
                field="DataFrame",
                expected="Данные",
                actual="Пусто",
                description="Входной DataFrame пуст или отсутствует"
            ))
            return self.issues

        logger.info(f"Запуск валидации для {len(df)} строк...")

        self._check_dates(df)
        self._check_missing_data(df)
        self._check_cost_discrepancies(df)
        self._check_volume_anomalies(df)
        self._check_price_anomalies(df)
        self._check_sku_consistency(df)
        self._check_duplicate_entries(df)

        logger.info(f"Валидация завершена: найдено {len(self.issues)} проблем")
        logger.info(f"  Критические: {sum(1 for i in self.issues if i.severity == SEVERITY_CRITICAL)}")
        logger.info(f"  Предупреждения: {sum(1 for i in self.issues if i.severity == SEVERITY_WARNING)}")
        logger.info(f"  Информация: {sum(1 for i in self.issues if i.severity == SEVERITY_INFO)}")

        return self.issues

    def _check_dates(self, df: pd.DataFrame):
        """Проверка дат: пустые, некорректные, логические ошибки."""
        date_cols = ['start_date', 'end_date', 'pdate']

        for col in date_cols:
            if col not in df.columns:
                continue

            # Пытаемся конвертировать в datetime
            dates = pd.to_datetime(df[col], dayfirst=True, errors='coerce')

            # Пустые/некорректные даты
            null_mask = dates.isna() & df[col].notna() & (df[col].astype(str).str.strip() != '')
            if null_mask.any():
                problem_rows = df[null_mask]
                for _, row in problem_rows.iterrows():
                    self.issues.append(ValidationIssue(
                        severity=SEVERITY_CRITICAL,
                        category="Даты",
                        file_name=str(row.get('FileName', '')),
                        sku=str(row.get('sku_type_sap', '')),
                        period=str(row.get('pdate', '')),
                        field=col,
                        expected="Корректная дата",
                        actual=str(row.get(col, '')),
                        description=f"Некорректный формат даты в поле '{col}'"
                    ))

            # Пустые даты
            empty_mask = dates.isna() & (df[col].isna() | (df[col].astype(str).str.strip() == ''))
            if empty_mask.any():
                count = empty_mask.sum()
                self.issues.append(ValidationIssue(
                    severity=SEVERITY_WARNING,
                    category="Даты",
                    file_name="(множество файлов)",
                    sku="",
                    period="",
                    field=col,
                    expected="Заполненная дата",
                    actual=f"Пусто ({count} строк)",
                    description=f"Отсутствует дата в поле '{col}' для {count} строк"
                ))

        # Дата начала > даты окончания
        if 'start_date' in df.columns and 'end_date' in df.columns:
            start_dates = pd.to_datetime(df['start_date'], dayfirst=True, errors='coerce')
            end_dates = pd.to_datetime(df['end_date'], dayfirst=True, errors='coerce')
            invalid_range = (start_dates > end_dates) & start_dates.notna() & end_dates.notna()

            if invalid_range.any():
                problem_rows = df[invalid_range]
                for _, row in problem_rows.head(10).iterrows():
                    self.issues.append(ValidationIssue(
                        severity=SEVERITY_CRITICAL,
                        category="Даты",
                        file_name=str(row.get('FileName', '')),
                        sku=str(row.get('sku_type_sap', '')),
                        period="",
                        field="start_date / end_date",
                        expected="start_date <= end_date",
                        actual=f"{row.get('start_date', '')} > {row.get('end_date', '')}",
                        description="Дата начала контракта позже даты окончания"
                    ))

        # pdate вне диапазона контракта
        if all(col in df.columns for col in ['pdate', 'start_date', 'end_date']):
            pdate = pd.to_datetime(df['pdate'], dayfirst=True, errors='coerce')
            start = pd.to_datetime(df['start_date'], dayfirst=True, errors='coerce')
            end = pd.to_datetime(df['end_date'], dayfirst=True, errors='coerce')

            out_of_range = ((pdate < start) | (pdate > end)) & pdate.notna() & start.notna() & end.notna()
            if out_of_range.any():
                count = out_of_range.sum()
                self.issues.append(ValidationIssue(
                    severity=SEVERITY_WARNING,
                    category="Даты",
                    file_name="(множество файлов)",
                    sku="",
                    period="",
                    field="pdate",
                    expected="В диапазоне контракта",
                    actual=f"{count} строк вне диапазона",
                    description=f"Период (pdate) вне диапазона контракта для {count} строк"
                ))

    def _check_missing_data(self, df: pd.DataFrame):
        """Проверка на пропущенные критичные данные."""
        critical_fields = {
            'viveska': 'Название на вывеске',
            'sku_type_sap': 'Тип SKU (SAP)',
            'sap-code': 'SAP-код',
            'filial': 'Филиал',
        }

        for field, display_name in critical_fields.items():
            if field not in df.columns:
                continue

            empty_mask = df[field].isna() | (df[field].astype(str).str.strip().isin(['', 'nan', 'None']))
            if empty_mask.any():
                count = empty_mask.sum()
                files = df[empty_mask]['FileName'].unique() if 'FileName' in df.columns else []
                self.issues.append(ValidationIssue(
                    severity=SEVERITY_WARNING if field != 'sku_type_sap' else SEVERITY_CRITICAL,
                    category="Пропущенные данные",
                    file_name=", ".join(str(f) for f in files[:5]) + ("..." if len(files) > 5 else ""),
                    sku="",
                    period="",
                    field=display_name,
                    expected="Заполнено",
                    actual=f"Пусто ({count} строк)",
                    description=f"Отсутствует '{display_name}' для {count} строк"
                ))

    def _check_cost_discrepancies(self, df: pd.DataFrame):
        """
        Проверка расхождений в затратах.
        Сравнивает плановые суммы с фактическими, выявляет аномальные расхождения.
        """
        cost_pairs = [
            ('Плановые затраты «Листинг/безусловные выплаты», руб',
             'Фактические затраты «Листинг/безусловные выплаты», руб',
             'Листинг'),
            ('Плановые затраты «Скидка в цене», руб',
             'Фактические затраты «Скидка в цене», руб',
             'Скидка в цене'),
            ('Плановые затраты «Ретро», руб',
             'Фактические затраты «Ретро», руб',
             'Ретро'),
            ('Плановые затраты «Маркетинг», руб',
             'Фактические затраты «Маркетинг», руб',
             'Маркетинг'),
            ('Плановые затраты «Промо-скидка», руб',
             'Фактические затраты «Промо-скидка», руб',
             'Промо-скидка'),
        ]

        for plan_col, fact_col, label in cost_pairs:
            if plan_col not in df.columns or fact_col not in df.columns:
                continue

            plan_vals = to_numeric_safe_with_null(df[plan_col])
            fact_vals = to_numeric_safe_with_null(df[fact_col])

            # Значительное расхождение (более чем в 5 раз)
            both_nonzero = (plan_vals > 0) & (fact_vals > 0)
            if both_nonzero.any():
                ratio = fact_vals[both_nonzero] / plan_vals[both_nonzero]
                extreme = (ratio > 5) | (ratio < 0.2)

                if extreme.any():
                    problem_rows = df[both_nonzero][extreme]
                    for _, row in problem_rows.head(10).iterrows():
                        plan_v = plan_vals[row.name]
                        fact_v = fact_vals[row.name]
                        self.issues.append(ValidationIssue(
                            severity=SEVERITY_CRITICAL,
                            category="Расхождение в затратах",
                            file_name=str(row.get('FileName', '')),
                            sku=str(row.get('sku_type_sap', '')),
                            period=str(row.get('pdate', '')),
                            field=label,
                            expected=f"План: {plan_v:,.0f}",
                            actual=f"Факт: {fact_v:,.0f}",
                            description=f"Расхождение план/факт по '{label}' "
                                        f"в {fact_v/plan_v:.1f} раз (план={plan_v:,.0f}, факт={fact_v:,.0f})"
                        ))

        # Общие затраты
        if 'план затраты' in df.columns and 'факт затраты' in df.columns:
            plan_total = to_numeric_safe_with_null(df['план затраты'])
            fact_total = to_numeric_safe_with_null(df['факт затраты'])

            both_nonzero = (plan_total > 0) & (fact_total > 0)
            if both_nonzero.any():
                ratio = fact_total[both_nonzero] / plan_total[both_nonzero]
                extreme = (ratio > 5) | (ratio < 0.2)
                if extreme.any():
                    count = extreme.sum()
                    self.issues.append(ValidationIssue(
                        severity=SEVERITY_CRITICAL,
                        category="Расхождение в затратах",
                        file_name="(множество файлов)",
                        sku="",
                        period="",
                        field="Общие затраты",
                        expected="План ≈ Факт (отклонение < 5x)",
                        actual=f"{count} строк с отклонением > 5x",
                        description=f"Итоговые затраты: {count} строк с расхождением план/факт более чем в 5 раз"
                    ))

    def _check_volume_anomalies(self, df: pd.DataFrame):
        """Проверка аномалий в объёмах."""
        volume_cols_map = {
            'volnew': 'Плановый объём',
            'Плановые продажи, шт': 'Плановые продажи',
            'Факт продажи, шт.': 'Фактические продажи',
        }

        for col, display_name in volume_cols_map.items():
            if col not in df.columns:
                continue

            vals = to_numeric_safe_with_null(df[col])

            # Отрицательные объёмы
            negative = vals < 0
            if negative.any():
                count = negative.sum()
                self.issues.append(ValidationIssue(
                    severity=SEVERITY_CRITICAL,
                    category="Аномалия объёмов",
                    file_name="(множество файлов)",
                    sku="",
                    period="",
                    field=display_name,
                    expected=">= 0",
                    actual=f"Отрицательные значения ({count} строк)",
                    description=f"Отрицательные значения в '{display_name}' для {count} строк"
                ))

            # Аномально большие объёмы (> 1 млн)
            huge = vals > 1_000_000
            if huge.any():
                count = huge.sum()
                self.issues.append(ValidationIssue(
                    severity=SEVERITY_INFO,
                    category="Аномалия объёмов",
                    file_name="(множество файлов)",
                    sku="",
                    period="",
                    field=display_name,
                    expected="< 1 000 000",
                    actual=f"> 1 000 000 ({count} строк)",
                    description=f"Аномально большие значения в '{display_name}' ({count} строк). Проверьте корректность."
                ))

    def _check_price_anomalies(self, df: pd.DataFrame):
        """Проверка аномалий в ценах."""
        if 'price' not in df.columns and 'price_in' not in df.columns:
            return

        for price_col, vol_col in [('price', 'volnew'), ('price_in', 'Плановые продажи, шт')]:
            if price_col not in df.columns or vol_col not in df.columns:
                continue

            prices = to_numeric_safe_with_null(df[price_col])
            volumes = to_numeric_safe_with_null(df[vol_col])

            # Нулевая цена при ненулевом объёме
            zero_price_with_vol = (prices == 0) & (volumes > 0)
            if zero_price_with_vol.any():
                count = zero_price_with_vol.sum()
                problem_rows = df[zero_price_with_vol]
                files = problem_rows['FileName'].unique() if 'FileName' in problem_rows.columns else []
                self.issues.append(ValidationIssue(
                    severity=SEVERITY_WARNING,
                    category="Аномалия цен",
                    file_name=", ".join(str(f) for f in files[:3]) + ("..." if len(files) > 3 else ""),
                    sku="",
                    period="",
                    field=price_col,
                    expected="Цена > 0 при объёме > 0",
                    actual=f"Цена = 0 при объёме > 0 ({count} строк)",
                    description=f"Нулевая цена при ненулевом объёме ({count} строк)"
                ))

    def _check_sku_consistency(self, df: pd.DataFrame):
        """Проверка консистентности SKU."""
        if 'sku_type_sap' not in df.columns:
            return

        # SKU без данных в конкретные месяцы (пропуски в последовательности)
        if 'FileName' in df.columns and 'pdate' in df.columns:
            for file_name in df['FileName'].unique():
                file_data = df[df['FileName'] == file_name]
                for sku in file_data['sku_type_sap'].unique():
                    sku_data = file_data[file_data['sku_type_sap'] == sku]
                    dates = pd.to_datetime(sku_data['pdate'], dayfirst=True, errors='coerce').dropna().sort_values()

                    if len(dates) >= 2:
                        # Проверяем на пропуски в месяцах
                        date_diffs = dates.diff().dt.days.dropna()
                        gaps = date_diffs[date_diffs > 45]  # Более 45 дней = пропущен месяц

                        if len(gaps) > 0:
                            self.issues.append(ValidationIssue(
                                severity=SEVERITY_WARNING,
                                category="Консистентность SKU",
                                file_name=str(file_name),
                                sku=str(sku),
                                period="",
                                field="pdate",
                                expected="Непрерывная последовательность месяцев",
                                actual=f"Пропуски в {len(gaps)} местах",
                                description=f"SKU '{sku}' имеет пропуски в последовательности месяцев"
                            ))

    def _check_duplicate_entries(self, df: pd.DataFrame):
        """Проверка на дубликаты."""
        key_cols = [c for c in ['FileName', 'viveska', 'sku_type_sap', 'pdate'] if c in df.columns]
        if len(key_cols) < 2:
            return

        duplicates = df.duplicated(subset=key_cols, keep=False)
        if duplicates.any():
            dup_count = duplicates.sum()
            dup_groups = df[duplicates].groupby(key_cols).size()
            multi_dups = dup_groups[dup_groups > 1]

            if len(multi_dups) > 0:
                self.issues.append(ValidationIssue(
                    severity=SEVERITY_WARNING,
                    category="Дубликаты",
                    file_name="(множество файлов)",
                    sku="",
                    period="",
                    field="Ключевые поля",
                    expected="Уникальные записи",
                    actual=f"{len(multi_dups)} групп дубликатов ({dup_count} строк)",
                    description=f"Обнаружены дублирующиеся записи по ключевым полям ({len(multi_dups)} групп)"
                ))


# =============================================================================
# ГЕНЕРАЦИЯ ОТЧЁТА О ВАЛИДАЦИИ
# =============================================================================

def generate_validation_report(issues: list[ValidationIssue], output_path: str):
    """Генерирует красивый Excel отчёт с результатами валидации."""
    wb = Workbook()

    # ====== Лист "Сводка" ======
    ws_summary = wb.active
    ws_summary.title = "Сводка валидации"

    # Стили
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
    title_font = Font(name='Calibri', size=16, bold=True, color="2F5496")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Заголовок
    ws_summary.merge_cells('A1:D1')
    ws_summary['A1'] = "ОТЧЁТ О ВАЛИДАЦИИ ДАННЫХ"
    ws_summary['A1'].font = title_font
    ws_summary['A1'].alignment = Alignment(horizontal='center')

    ws_summary['A3'] = f"Дата проверки: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}"
    ws_summary['A3'].font = Font(italic=True)

    # Статистика по серьёзности
    ws_summary['A5'] = "Категория"
    ws_summary['B5'] = "Количество"
    ws_summary['A5'].font = Font(bold=True)
    ws_summary['B5'].font = Font(bold=True)

    critical_count = sum(1 for i in issues if i.severity == SEVERITY_CRITICAL)
    warning_count = sum(1 for i in issues if i.severity == SEVERITY_WARNING)
    info_count = sum(1 for i in issues if i.severity == SEVERITY_INFO)

    for row_idx, (label, count, color) in enumerate([
        (SEVERITY_CRITICAL, critical_count, "FF4444"),
        (SEVERITY_WARNING, warning_count, "FFAA00"),
        (SEVERITY_INFO, info_count, "4488FF"),
    ], start=6):
        ws_summary.cell(row=row_idx, column=1, value=label).fill = PatternFill(
            start_color=color, end_color=color, fill_type="solid"
        )
        ws_summary.cell(row=row_idx, column=1).font = Font(bold=True, color="FFFFFF")
        ws_summary.cell(row=row_idx, column=2, value=count)

    ws_summary['A9'] = f"Всего проблем: {len(issues)}"
    ws_summary['A9'].font = Font(bold=True, size=12)

    # Статистика по категориям
    ws_summary['A11'] = "По категориям:"
    ws_summary['A11'].font = Font(bold=True)

    categories = {}
    for issue in issues:
        categories[issue.category] = categories.get(issue.category, 0) + 1

    for row_idx, (cat, count) in enumerate(sorted(categories.items()), start=12):
        ws_summary.cell(row=row_idx, column=1, value=cat)
        ws_summary.cell(row=row_idx, column=2, value=count)

    ws_summary.column_dimensions['A'].width = 40
    ws_summary.column_dimensions['B'].width = 15

    # ====== Лист "Детали" ======
    ws_details = wb.create_sheet("Детали проблем")

    # Заголовки
    headers = ['Серьёзность', 'Категория', 'Файл', 'SKU', 'Период', 'Поле', 'Ожидалось', 'Фактически', 'Описание']
    for col_idx, header in enumerate(headers, 1):
        cell = ws_details.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # Данные
    for row_idx, issue in enumerate(issues, 2):
        data = issue.to_dict()
        for col_idx, key in enumerate(headers, 1):
            cell = ws_details.cell(row=row_idx, column=col_idx, value=data.get(key, ''))
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)

            # Цветовая маркировка серьёзности
            if col_idx == 1:
                color = SEVERITY_COLORS.get(data.get('Серьёзность', ''), "FFFFFF")
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")

    # Автоширина колонок
    col_widths = [18, 25, 30, 25, 15, 25, 25, 25, 50]
    for col_idx, width in enumerate(col_widths, 1):
        ws_details.column_dimensions[get_column_letter(col_idx)].width = width

    ws_details.freeze_panes = 'A2'

    # Сохраняем
    wb.save(output_path)
    logger.info(f"Отчёт валидации сохранён: {output_path}")

    return output_path


def validate_data(df: pd.DataFrame, output_path: str = None) -> tuple:
    """
    Выполняет полную валидацию данных и опционально сохраняет отчёт.

    Args:
        df: DataFrame для проверки
        output_path: путь для сохранения отчёта (опционально)

    Returns:
        tuple: (список проблем, путь к отчёту или None)
    """
    validator = DataValidator()
    issues = validator.validate_all(df)

    report_path = None
    if output_path and issues:
        report_path = generate_validation_report(issues, output_path)

    return issues, report_path
