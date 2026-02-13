"""
Шаг 4: Генерация итогового Excel-отчёта.

Логика из оригинального кода "4.py":
- Загрузка и объединение данных
- Расчёт P&L (план/факт продаж, затраты, доход)
- Поддержка дополнительных источников (продажи, затраты, ЦМ, себестоимость)
- Вывод в красиво оформленный Excel файл
"""
import os
import logging

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from utils import to_numeric_safe_with_null

logger = logging.getLogger(__name__)


# =============================================================================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# =============================================================================

def save_to_excel_with_chunks(df, output_path, chunk_size=800000):
    """Сохраняет DataFrame в Excel с разбиением на листы при необходимости."""
    n_chunks = len(df) // chunk_size + (1 if len(df) % chunk_size != 0 else 0)
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for i in range(n_chunks):
            start_idx = i * chunk_size
            end_idx = min((i + 1) * chunk_size, len(df))
            chunk = df.iloc[start_idx:end_idx].copy()
            sheet_name = f'Sheet{i+1}' if n_chunks > 1 else 'Sheet1'
            chunk.to_excel(writer, sheet_name=sheet_name, index=False)
            logger.info(f"Сохранено {len(chunk)} строк в лист '{sheet_name}'")
    logger.info(f"Файл сохранён: {output_path}")


def apply_excel_formatting(output_path):
    """Применяет красивое форматирование к Excel файлу."""
    from openpyxl import load_workbook

    wb = load_workbook(output_path)

    # Цвета
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(name='Calibri', size=10, bold=True, color="FFFFFF")
    plan_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    fact_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    diff_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    warning_font = Font(name='Calibri', size=10, color="FF0000", bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for ws in wb.worksheets:
        if ws.max_row <= 1:
            continue

        # Форматируем заголовки
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

        # Форматируем данные
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')

                header_value = str(ws.cell(row=1, column=col).value or "").lower()

                # Цветовая маркировка по типу колонки
                if 'плановые' in header_value or 'план' in header_value:
                    cell.fill = plan_fill
                elif 'фактические' in header_value or 'факт' in header_value:
                    cell.fill = fact_fill
                elif 'разница' in header_value:
                    cell.fill = diff_fill
                    # Красный цвет для отрицательных значений
                    try:
                        if cell.value is not None and float(cell.value) < 0:
                            cell.font = warning_font
                    except (ValueError, TypeError):
                        pass

        # Автоподбор ширины
        for col in range(1, ws.max_column + 1):
            max_length = 0
            col_letter = get_column_letter(col)
            for row in range(1, min(ws.max_row + 1, 100)):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    try:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                    except Exception:
                        pass
            ws.column_dimensions[col_letter].width = min(max_length + 4, 40)

        # Фиксируем первую строку
        ws.freeze_panes = 'A2'

    wb.save(output_path)
    logger.info(f"Форматирование применено к {output_path}")


# =============================================================================
# ГЕНЕРАЦИЯ P&L ОТЧЁТА
# =============================================================================

def generate_pnl_report(
    merged_df: pd.DataFrame,
    output_path: str,
    sales_file: str = None,
    costs_not_price_file: str = None,
    costs_in_price_file: str = None,
    cm_file: str = None,
    cogs_file: str = None,
    methodichka_file: str = None
) -> str:
    """
    Генерирует P&L отчёт из объединённых данных.

    Args:
        merged_df: объединённый DataFrame с данными контрактов
        output_path: путь для сохранения итогового файла
        sales_file: путь к файлу с фактическими продажами (опционально)
        costs_not_price_file: путь к файлу затрат вне цены (опционально)
        costs_in_price_file: путь к файлу затрат в цене (опционально)
        cm_file: путь к файлу ЦМ (опционально)
        cogs_file: путь к файлу себестоимости (опционально)
        methodichka_file: путь к файлу методички для нормализации SKU (опционально)

    Returns:
        Путь к сохранённому файлу
    """
    logger.info("Генерация P&L отчёта...")

    if merged_df.empty:
        logger.warning("Нет данных для отчёта")
        return None

    # Копия для безопасной работы
    df = merged_df.copy()

    # Нормализуем числовые колонки
    numeric_cols = [
        'price', 'price_in', 'listing', 'listing2',
        'marketing', 'marketing2', 'promo', 'promo2',
        'retro', 'volnew', 'PromVol', 'dopmarketing'
    ]
    for col in numeric_cols:
        if col in df.columns:
            df[col] = to_numeric_safe_with_null(df[col])

    # Даты
    for date_col in ['pdate', 'start_date', 'end_date']:
        if date_col in df.columns:
            df[date_col] = pd.to_datetime(df[date_col], errors='coerce')

    # Группируем плановые данные
    group_cols = ['FileName', 'viveska', 'gr_sb', 'sku_type_sap', 'pdate', 'start_date', 'end_date']
    available_group_cols = [c for c in group_cols if c in df.columns]

    if 'volnew' in df.columns:
        df_plan = df.groupby(available_group_cols, as_index=False)['volnew'].sum()
        df_plan.rename(columns={'volnew': 'Плановые продажи, шт'}, inplace=True)
    else:
        df_plan = df[available_group_cols].drop_duplicates()
        df_plan['Плановые продажи, шт'] = 0

    df_combined = df_plan.copy()

    # Добавляем ЦМ если файл указан
    if cm_file and os.path.exists(cm_file):
        try:
            df_cm = pd.read_excel(cm_file)
            if 'ЦМ' in df_cm.columns:
                df_cm['ЦМ'] = to_numeric_safe_with_null(df_cm['ЦМ'])
                merge_cols = [c for c in ['gr_sb', 'sku_type_sap', 'pdate'] if c in df_cm.columns and c in df_combined.columns]
                if merge_cols:
                    df_combined = pd.merge(df_combined, df_cm[merge_cols + ['ЦМ']], on=merge_cols, how='left')
                    df_combined.rename(columns={'ЦМ': 'price_in'}, inplace=True)
        except Exception as e:
            logger.warning(f"Ошибка загрузки ЦМ: {e}")

    if 'price_in' not in df_combined.columns:
        # Берём price_in из исходных данных
        if 'price_in' in df.columns:
            price_in_data = df.groupby(available_group_cols, as_index=False)['price_in'].first()
            df_combined = pd.merge(df_combined, price_in_data, on=available_group_cols, how='left')

    if 'price_in' in df_combined.columns:
        df_combined['price_in'] = to_numeric_safe_with_null(df_combined['price_in'])
    else:
        df_combined['price_in'] = 0.0

    df_combined['Плановые продажи, руб'] = df_combined['Плановые продажи, шт'] * df_combined['price_in']

    # Добавляем инвестиционные показатели
    for col in ['listing2', 'listing', 'retro', 'dopmarketing', 'marketing', 'PromVol']:
        if col in df.columns:
            grouped = df.groupby(available_group_cols, as_index=False)[col].sum()
            df_combined = pd.merge(df_combined, grouped, on=available_group_cols, how='left')
            df_combined[col] = to_numeric_safe_with_null(df_combined[col])
        else:
            df_combined[col] = 0.0

    # Расчёт плановых затрат
    df_combined['Плановые затраты «Скидка в цене», руб'] = \
        df_combined['Плановые продажи, руб'] * df_combined['listing2']

    df_combined['Плановые затраты «Листинг/безусловные выплаты», руб'] = df_combined['listing']

    df_combined['Плановые затраты «Ретро», руб'] = \
        (df_combined['Плановые продажи, руб'] / 1.2) * df_combined['retro']

    df_combined['Плановые затраты «Маркетинг», руб'] = \
        df_combined['Плановые продажи, руб'] * df_combined['dopmarketing'] + df_combined['marketing']

    df_combined['Плановые затраты «Промо-скидка», руб'] = df_combined['PromVol']

    # Статус контракта
    if 'pdate' in df_combined.columns and 'end_date' in df_combined.columns:
        df_combined['контракт'] = np.where(
            pd.to_datetime(df_combined['pdate']) > pd.to_datetime(df_combined['end_date']),
            'завершенный', 'действующий'
        )
    else:
        df_combined['контракт'] = 'действующий'

    # Фактические продажи (если файл указан)
    if sales_file and os.path.exists(sales_file):
        try:
            df_sales = pd.read_excel(sales_file)
            df_sales['sales_date'] = pd.to_datetime(df_sales.get('sales_date', pd.Series()), errors='coerce')
            if 'vol_2' in df_sales.columns:
                df_sales['vol_2'] = to_numeric_safe_with_null(df_sales['vol_2'])
                merge_cols = [c for c in ['viveska', 'sku_type_sap', 'pdate'] if c in df_sales.columns]
                if merge_cols:
                    sales_agg = df_sales.groupby(merge_cols, as_index=False)['vol_2'].sum()
                    df_combined = pd.merge(df_combined, sales_agg, on=merge_cols, how='left')
                    df_combined.rename(columns={'vol_2': 'Факт продажи, шт.'}, inplace=True)
        except Exception as e:
            logger.warning(f"Ошибка загрузки продаж: {e}")

    if 'Факт продажи, шт.' not in df_combined.columns:
        df_combined['Факт продажи, шт.'] = 0.0

    df_combined['Факт продажи, шт.'] = to_numeric_safe_with_null(df_combined['Факт продажи, шт.'])
    df_combined['Факт продажи, руб (от ЦМ)'] = df_combined['Факт продажи, шт.'] * df_combined['price_in']
    df_combined['Разница, шт'] = df_combined['Факт продажи, шт.'] - df_combined['Плановые продажи, шт']
    df_combined['Разница, руб'] = df_combined['Факт продажи, руб (от ЦМ)'] - df_combined['Плановые продажи, руб']

    # Фактические затраты (если файлы указаны)
    fact_expense_cols = [
        'Фактические затраты «Листинг/безусловные выплаты», руб',
        'Фактические затраты «Ретро», руб',
        'Фактические затраты «Маркетинг», руб',
        'Фактические затраты «Промо-скидка», руб',
        'Фактические затраты «Скидка в цене», руб',
    ]

    if costs_not_price_file and os.path.exists(costs_not_price_file):
        try:
            df_cnp = pd.read_excel(costs_not_price_file)
            if 'Сумма' in df_cnp.columns and 'Статья расходов' in df_cnp.columns:
                df_cnp['Сумма'] = to_numeric_safe_with_null(df_cnp['Сумма'])
                df_cnp['pdate'] = pd.to_datetime(df_cnp.get('Месяц/год', pd.Series()), errors='coerce')

                for expense, col_name in [
                    ('Листинг', 'Фактические затраты «Листинг/безусловные выплаты», руб'),
                    ('Маркетинг', 'Фактические затраты «Маркетинг», руб'),
                    ('Ретро', 'Фактические затраты «Ретро», руб')
                ]:
                    filtered = df_cnp[df_cnp['Статья расходов'].str.contains(expense, case=False, na=False)]
                    if not filtered.empty:
                        merge_cols = [c for c in ['pdate', 'viveska', 'sku_type_sap'] if c in filtered.columns and c in df_combined.columns]
                        if merge_cols:
                            grouped = filtered.groupby(merge_cols, as_index=False)['Сумма'].sum()
                            df_combined = pd.merge(df_combined, grouped, on=merge_cols, how='left')
                            df_combined.rename(columns={'Сумма': col_name}, inplace=True)
        except Exception as e:
            logger.warning(f"Ошибка загрузки затрат вне цены: {e}")

    if costs_in_price_file and os.path.exists(costs_in_price_file):
        try:
            df_cip = pd.read_excel(costs_in_price_file)
            if 'Сумма в валюте документа' in df_cip.columns:
                df_cip['Сумма в валюте документа'] = to_numeric_safe_with_null(df_cip['Сумма в валюте документа'])
                df_cip['pdate'] = pd.to_datetime(df_cip.get('Месяц/год', pd.Series()), errors='coerce')

                if 'примечание' in df_cip.columns:
                    merge_cols = [c for c in ['pdate', 'viveska', 'sku_type_sap'] if c in df_cip.columns and c in df_combined.columns]
                    if merge_cols:
                        # Промо-скидка
                        promo = df_cip[df_cip['примечание'].str.contains('промо акция', case=False, na=False)]
                        if not promo.empty:
                            promo_group = promo.groupby(merge_cols, as_index=False)['Сумма в валюте документа'].sum()
                            df_combined = pd.merge(df_combined, promo_group, on=merge_cols, how='left')
                            df_combined.rename(columns={'Сумма в валюте документа': 'Фактические затраты «Промо-скидка», руб'}, inplace=True)

                        # Скидка в цене
                        skidka = df_cip[df_cip['примечание'].str.contains('скидка в цене', case=False, na=False)]
                        if not skidka.empty:
                            skidka_group = skidka.groupby(merge_cols, as_index=False)['Сумма в валюте документа'].sum()
                            df_combined = pd.merge(df_combined, skidka_group, on=merge_cols, how='left')
                            df_combined.rename(columns={'Сумма в валюте документа': 'Фактические затраты «Скидка в цене», руб'}, inplace=True)
        except Exception as e:
            logger.warning(f"Ошибка загрузки затрат в цене: {e}")

    # Заполняем нулями отсутствующие колонки затрат
    for col in fact_expense_cols:
        if col not in df_combined.columns:
            df_combined[col] = 0.0
        else:
            df_combined[col] = to_numeric_safe_with_null(df_combined[col])

    # Итоговые затраты
    df_combined['план затраты'] = (
        df_combined.get('Плановые затраты «Листинг/безусловные выплаты», руб', 0) +
        df_combined.get('Плановые затраты «Скидка в цене», руб', 0) +
        df_combined.get('Плановые затраты «Ретро», руб', 0) +
        df_combined.get('Плановые затраты «Маркетинг», руб', 0) +
        df_combined.get('Плановые затраты «Промо-скидка», руб', 0)
    )

    df_combined['факт затраты'] = (
        df_combined.get('Фактические затраты «Листинг/безусловные выплаты», руб', 0) +
        df_combined.get('Фактические затраты «Скидка в цене», руб', 0) +
        df_combined.get('Фактические затраты «Ретро», руб', 0) +
        df_combined.get('Фактические затраты «Маркетинг», руб', 0) +
        df_combined.get('Фактические затраты «Промо-скидка», руб', 0)
    )

    # Себестоимость (если файл указан)
    if cogs_file and os.path.exists(cogs_file):
        try:
            df_cogs = pd.read_excel(cogs_file)
            if 'cogs' in df_cogs.columns:
                df_cogs['cogs'] = to_numeric_safe_with_null(df_cogs['cogs'])
                merge_cols = [c for c in ['sku_type_sap', 'pdate'] if c in df_cogs.columns and c in df_combined.columns]
                if merge_cols:
                    df_combined = pd.merge(df_combined, df_cogs[merge_cols + ['cogs']], on=merge_cols, how='left')
        except Exception as e:
            logger.warning(f"Ошибка загрузки себестоимости: {e}")

    if 'cogs' not in df_combined.columns:
        df_combined['cogs'] = 0.0
    df_combined['cogs'] = to_numeric_safe_with_null(df_combined['cogs'])

    df_combined['продажи по сс план'] = df_combined['Плановые продажи, шт'] * df_combined['cogs']
    df_combined['продажи по сс факт'] = df_combined['Факт продажи, шт.'] * df_combined['cogs']

    df_combined['доход план'] = df_combined['Плановые продажи, руб'] - df_combined['продажи по сс план'] - df_combined['план затраты']
    df_combined['доход факт'] = df_combined['Факт продажи, руб (от ЦМ)'] - df_combined['продажи по сс факт'] - df_combined['факт затраты']

    # Удаление дубликатов
    df_combined = df_combined.drop_duplicates(keep='first')
    key_columns = [c for c in ['pdate', 'FileName', 'viveska', 'gr_sb', 'sku_type_sap', 'start_date', 'end_date']
                    if c in df_combined.columns]
    if key_columns:
        df_combined = df_combined.drop_duplicates(subset=key_columns, keep='first')

    # Порядок колонок
    desired_order = [
        'pdate', 'FileName', 'viveska', 'gr_sb', 'sku_type_sap',
        'start_date', 'end_date', 'контракт',
        'Плановые продажи, шт', 'Факт продажи, шт.', 'Разница, шт',
        'Плановые продажи, руб', 'Факт продажи, руб (от ЦМ)', 'Разница, руб',
        'Плановые затраты «Листинг/безусловные выплаты», руб',
        'Фактические затраты «Листинг/безусловные выплаты», руб',
        'Плановые затраты «Скидка в цене», руб',
        'Фактические затраты «Скидка в цене», руб',
        'Плановые затраты «Ретро», руб',
        'Фактические затраты «Ретро», руб',
        'Плановые затраты «Маркетинг», руб',
        'Фактические затраты «Маркетинг», руб',
        'Плановые затраты «Промо-скидка», руб',
        'Фактические затраты «Промо-скидка», руб',
        'план затраты', 'факт затраты',
        'продажи по сс план', 'продажи по сс факт',
        'доход план', 'доход факт'
    ]
    desired_order = [col for col in desired_order if col in df_combined.columns]
    # Добавляем остальные колонки, которых нет в порядке
    for col in df_combined.columns:
        if col not in desired_order:
            desired_order.append(col)
    df_combined = df_combined[desired_order]

    # Сохраняем
    save_to_excel_with_chunks(df_combined, output_path)

    # Применяем красивое форматирование
    try:
        apply_excel_formatting(output_path)
    except Exception as e:
        logger.warning(f"Не удалось применить форматирование: {e}")

    logger.info(f"P&L отчёт сохранён: {output_path}")
    return output_path


def generate_simple_report(merged_df: pd.DataFrame, output_path: str) -> str:
    """
    Генерирует простой отчёт без дополнительных источников данных.
    Только на основе данных из обработанных файлов контрактов.
    """
    return generate_pnl_report(merged_df, output_path)
