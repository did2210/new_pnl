import openpyxl
import os
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.styles.colors import Color
from datetime import datetime
import traceback
import time
from tqdm import tqdm  # Для красивого прогресс-бара

def convert_to_string(value):
    """Преобразует любое значение в строку, обрабатывая кортежи и другие сложные типы"""
    if value is None:
        return ""
    # Обработка кортежей
    if isinstance(value, tuple):
        # Если это кортеж с одним элементом, возвращаем его как строку
        if len(value) == 1:
            return str(value[0])
        # Иначе преобразуем весь кортеж в строку
        return ", ".join([str(item) for item in value])
    # Обработка других типов
    return str(value)

def find_text_in_excel(file_path, search_text, data_only=True):
    """
    Ищет указанный текст в Excel файле и возвращает координаты ячейки
    Args:
        file_path (str): Путь к Excel файлу
        search_text (str): Текст для поиска
        data_only (bool): Если True, возвращает рассчитанные значения, а не формулы
    Returns:
        tuple: (номер строки, номер столбца, имя листа, ячейка, рабочая книга) или None
    """
    try:
        # Загружаем рабочую книгу с data_only=True для получения рассчитанных значений
        workbook = openpyxl.load_workbook(file_path, data_only=data_only)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row_idx, row in enumerate(sheet.iter_rows(values_only=False), 1):
                for col_idx, cell in enumerate(row, 1):
                    # Проверяем значение ячейки
                    cell_value = cell.value
                    # Преобразуем значение в строку для поиска
                    if cell_value is not None:
                        cell_value_str = convert_to_string(cell_value)
                        if isinstance(cell_value_str, str) and search_text in cell_value_str:
                            return (row_idx, col_idx, sheet_name, cell, workbook)
        return None
    except Exception as e:
        return None

def extract_gfd_request_table(file_path, output_dir):
    """
    Ищет и извлекает таблицу "Запрос на заключение контракта по напиткам GFD"
    """
    try:
        # Поиск ячейки с текстом "Запрос на заключение контракта по напиткам GFD"
        result = find_text_in_excel(file_path, "Запрос на заключение контракта по напиткам GFD")
        if not result:
            print("Не удалось найти текст 'Запрос на заключение контракта по напиткам GFD'")
            return None, 0
        row, col, sheet_name, cell, workbook = result
        sheet = workbook[sheet_name]
        print(f"Найден заголовок 'Запрос на заключение контракта по напиткам GFD' в ячейке {get_column_letter(col)}{row}")
        # Поиск конца таблицы (начало следующего раздела)
        end_row = None
        for r in range(row + 1, sheet.max_row + 1):
            for c in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=r, column=c)
                if cell.value and isinstance(cell.value, str):
                    # Ищем начало следующего раздела
                    if "Условия для нового контракта" in cell.value or "Условия контракта" in cell.value:
                        end_row = r - 1
                        print(f"Найден конец таблицы перед '{cell.value}' в строке {r}")
                        break
            if end_row:
                break
        # Если не найден явный конец, используем эвристику для определения конца таблицы
        if not end_row:
            # Ищем пустую строку после заголовка
            empty_row_count = 0
            for r in range(row + 1, sheet.max_row + 1):
                is_empty = True
                for c in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=r, column=c)
                    if cell.value is not None:
                        is_empty = False
                        break
                if is_empty:
                    empty_row_count += 1
                    if empty_row_count >= 2:  # Две пустые строки подряд
                        end_row = r - 2
                        break
                else:
                    empty_row_count = 0
            # Если не нашли пустые строки, используем максимальную строку
            if not end_row:
                end_row = sheet.max_row
        # Определяем начало данных (пропускаем заголовок)
        data_start_row = row + 1
        print(f"Извлечение данных GFD таблицы с {data_start_row} по {end_row} строку...")
        # Создаем новый рабочий лист
        excel_wb = openpyxl.Workbook()
        excel_ws = excel_wb.active
        excel_ws.title = "GFD Запрос"
        # Добавляем заголовок
        excel_ws.merge_cells('A1:Z1')
        header_cell = excel_ws['A1']
        header_cell.value = "ЗАПРОС НА ЗАКЛЮЧЕНИЕ КОНТРАКТА ПО НАПИТКАМ GFD"
        header_cell.font = Font(bold=True, size=16)
        header_cell.alignment = Alignment(horizontal='center')
        excel_ws.merge_cells('A2:Z2')
        excel_ws['A2'] = f"Дата извлечения: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Лист: {sheet_name} | Диапазон: с {get_column_letter(col)}{data_start_row} по {get_column_letter(sheet.max_column)}{end_row}"
        excel_ws['A2'].font = Font(italic=True)
        excel_ws['A2'].alignment = Alignment(horizontal='center')
        # Добавляем отступ
        current_excel_row = 4
        # Копируем данные в новый файл
        for r in range(data_start_row, end_row + 1):
            for c in range(1, sheet.max_column + 1):
                source_cell = sheet.cell(row=r, column=c)
                target_cell = excel_ws.cell(row=current_excel_row, column=c)
                # Копируем значение
                if source_cell.value is not None:
                    target_cell.value = source_cell.value
                # Копируем стиль (безопасный способ)
                if source_cell.has_style:
                    # Копируем шрифт
                    if source_cell.font:
                        try:
                            target_cell.font = Font(
                                name=source_cell.font.name,
                                size=source_cell.font.size,
                                bold=source_cell.font.bold,
                                italic=source_cell.font.italic,
                                underline=source_cell.font.underline,
                                color=source_cell.font.color
                            )
                        except:
                            target_cell.font = Font()
                    # Копируем границы
                    if source_cell.border:
                        try:
                            target_cell.border = Border(
                                left=source_cell.border.left,
                                right=source_cell.border.right,
                                top=source_cell.border.top,
                                bottom=source_cell.border.bottom
                            )
                        except:
                            target_cell.border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin')
                            )
                    # Копируем заливку
                    if source_cell.fill:
                        try:
                            if source_cell.fill.fill_type == "solid" and source_cell.fill.fgColor:
                                try:
                                    rgb = source_cell.fill.fgColor.rgb
                                    target_cell.fill = PatternFill(
                                        fill_type="solid",
                                        fgColor=rgb
                                    )
                                except:
                                    target_cell.fill = PatternFill(
                                        fill_type=source_cell.fill.fill_type,
                                        fgColor="FFFFFF"
                                    )
                            else:
                                target_cell.fill = source_cell.fill
                        except:
                            target_cell.fill = PatternFill(fill_type=None)
                    # Копируем выравнивание
                    if source_cell.alignment:
                        try:
                            target_cell.alignment = Alignment(
                                horizontal=source_cell.alignment.horizontal,
                                vertical=source_cell.alignment.vertical,
                                wrap_text=source_cell.alignment.wrap_text,
                                indent=source_cell.alignment.indent
                            )
                        except:
                            target_cell.alignment = Alignment(horizontal='left', vertical='center')
                    # Копируем числовой формат
                    if hasattr(source_cell, 'number_format') and source_cell.number_format:
                        try:
                            target_cell.number_format = source_cell.number_format
                        except:
                            target_cell.number_format = 'General'
            current_excel_row += 1
        # Добавляем рамку вокруг данных
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        for r in range(4, current_excel_row):
            for c in range(1, sheet.max_column + 1):
                cell = excel_ws.cell(row=r, column=c)
                # Если у ячейки нет границ, добавляем рамку
                if not cell.border or (not cell.border.left.style and not cell.border.right.style and 
                                      not cell.border.top.style and not cell.border.bottom.style):
                    cell.border = thin_border
        # Автоподбор ширины столбцов
        for col in range(1, sheet.max_column + 1):
            max_length = 0
            column = get_column_letter(col)
            for row in range(4, current_excel_row):
                cell = excel_ws.cell(row=row, column=col)
                if cell.value:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            adjusted_width = (max_length + 2)
            if adjusted_width > 50:
                adjusted_width = 50
            excel_ws.column_dimensions[column].width = adjusted_width
        # Добавляем подвал
        excel_ws.merge_cells(f'A{current_excel_row}:Z{current_excel_row}')
        excel_ws[f'A{current_excel_row}'] = f"Отчет сгенерирован автоматически {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        excel_ws[f'A{current_excel_row}'].font = Font(italic=True, size=10)
        excel_ws[f'A{current_excel_row}'].alignment = Alignment(horizontal='right')
        return excel_wb, current_excel_row - 4
    except Exception as e:
        print(f"Произошла ошибка при извлечении GFD таблицы: {e}")
        traceback.print_exc()
        return None, 0

def extract_contract_conditions_table(file_path, output_dir):
    """
    Ищет и извлекает таблицу условий контракта
    """
    try:
        # Поиск ячейки с текстом "Условия для нового контракта"
        result = find_text_in_excel(file_path, "Условия для нового контракта", data_only=True)
        if not result:
            print("Не удалось найти текст 'Условия для нового контракта'")
            return None, 0
        row, col, sheet_name, cell, workbook = result
        sheet = workbook[sheet_name]
        # Поиск ячейки с текстом "ВСЕГО:"
        total_cell = None
        for r in range(row, sheet.max_row + 1):
            for c in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=r, column=c)
                if cell.value is not None:
                    cell_value_str = convert_to_string(cell.value)
                    if isinstance(cell_value_str, str) and ("ВСЕГО:" in cell_value_str or "ВСЕГО (" in cell_value_str):
                        total_cell = (r, c, cell)
                        break
            if total_cell:
                break
        if not total_cell:
            print("Ячейка 'ВСЕГО:' не найдена после ячейки с условиями контракта")
            return None, 0
        # Извлечение данных между найденными ячейками
        start_row = row + 1  # Начинаем с ячейки ниже "Условия для нового контракта"
        end_row = total_cell[0] - 1  # Заканчиваем ячейкой выше "ВСЕГО:"
        print(f"Извлечение данных условий контракта с {start_row} по {end_row} строку...")
        # Создаем новый рабочий лист
        excel_wb = openpyxl.Workbook()
        excel_ws = excel_wb.active
        excel_ws.title = "Условия контракта"
        # Добавляем заголовок
        excel_ws.merge_cells('A1:Z1')
        header_cell = excel_ws['A1']
        header_cell.value = "УСЛОВИЯ КОНТРАКТА"
        header_cell.font = Font(bold=True, size=16)
        header_cell.alignment = Alignment(horizontal='center')
        excel_ws.merge_cells('A2:Z2')
        excel_ws['A2'] = f"Дата извлечения: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Лист: {sheet_name} | Диапазон: с {get_column_letter(col)}{start_row} по {get_column_letter(total_cell[1])}{end_row}"
        excel_ws['A2'].font = Font(italic=True)
        excel_ws['A2'].alignment = Alignment(horizontal='center')
        # Добавляем отступ
        current_excel_row = 4
        # Копируем данные в новый файл
        for r in range(start_row, end_row + 1):
            for c in range(1, sheet.max_column + 1):
                source_cell = sheet.cell(row=r, column=c)
                target_cell = excel_ws.cell(row=current_excel_row, column=c)
                # Копируем значение
                if source_cell.value is not None:
                    target_cell.value = source_cell.value
                # Копируем стиль (безопасный способ)
                if source_cell.has_style:
                    # Копируем шрифт
                    if source_cell.font:
                        try:
                            target_cell.font = Font(
                                name=source_cell.font.name,
                                size=source_cell.font.size,
                                bold=source_cell.font.bold,
                                italic=source_cell.font.italic,
                                underline=source_cell.font.underline,
                                color=source_cell.font.color
                            )
                        except:
                            target_cell.font = Font()
                    # Копируем границы
                    if source_cell.border:
                        try:
                            target_cell.border = Border(
                                left=source_cell.border.left,
                                right=source_cell.border.right,
                                top=source_cell.border.top,
                                bottom=source_cell.border.bottom
                            )
                        except:
                            target_cell.border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin')
                            )
                    # Копируем заливку
                    if source_cell.fill:
                        try:
                            if source_cell.fill.fill_type == "solid" and source_cell.fill.fgColor:
                                try:
                                    rgb = source_cell.fill.fgColor.rgb
                                    target_cell.fill = PatternFill(
                                        fill_type="solid",
                                        fgColor=rgb
                                    )
                                except:
                                    target_cell.fill = PatternFill(
                                        fill_type=source_cell.fill.fill_type,
                                        fgColor="FFFFFF"
                                    )
                            else:
                                target_cell.fill = source_cell.fill
                        except:
                            target_cell.fill = PatternFill(fill_type=None)
                    # Копируем выравнивание
                    if source_cell.alignment:
                        try:
                            target_cell.alignment = Alignment(
                                horizontal=source_cell.alignment.horizontal,
                                vertical=source_cell.alignment.vertical,
                                wrap_text=source_cell.alignment.wrap_text,
                                indent=source_cell.alignment.indent
                            )
                        except:
                            target_cell.alignment = Alignment(horizontal='left', vertical='center')
                    # Копируем числовой формат
                    if hasattr(source_cell, 'number_format') and source_cell.number_format:
                        try:
                            target_cell.number_format = source_cell.number_format
                        except:
                            target_cell.number_format = 'General'
            current_excel_row += 1
        # Добавляем рамку вокруг данных
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        for r in range(4, current_excel_row):
            for c in range(1, sheet.max_column + 1):
                cell = excel_ws.cell(row=r, column=c)
                # Если у ячейки нет границ, добавляем рамку
                if not cell.border or (not cell.border.left.style and not cell.border.right.style and 
                                      not cell.border.top.style and not cell.border.bottom.style):
                    cell.border = thin_border
        # Автоподбор ширины столбцов
        for col in range(1, sheet.max_column + 1):
            max_length = 0
            column = get_column_letter(col)
            for row in range(4, current_excel_row):
                cell = excel_ws.cell(row=row, column=col)
                if cell.value:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            adjusted_width = (max_length + 2)
            if adjusted_width > 50:
                adjusted_width = 50
            excel_ws.column_dimensions[column].width = adjusted_width
        # Добавляем подвал
        excel_ws.merge_cells(f'A{current_excel_row}:Z{current_excel_row}')
        excel_ws[f'A{current_excel_row}'] = f"Отчет сгенерирован автоматически {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        excel_ws[f'A{current_excel_row}'].font = Font(italic=True, size=10)
        excel_ws[f'A{current_excel_row}'].alignment = Alignment(horizontal='right')
        return excel_wb, current_excel_row - 4
    except Exception as e:
        print(f"Произошла ошибка при извлечении таблицы условий контракта: {e}")
        traceback.print_exc()
        return None, 0

def extract_planning_sales_data(file_path, output_dir):
    """
    Ищет и извлекает данные между "Блок ПЛАНИРОВАНИЕ продаж" и
    началом следующей секции или концом листа.
    """
    try:
        print("--- Поиск данных планирования продаж ---")
        # Загружаем рабочую книгу
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        # Ищем лист с нужными данными (предположим, что это лист "NEW CNR 1", "Расчет инвестиций", "NEW CNR" или "Расчет инвестиций (2)")
        target_sheet_names = ["NEW CNR 1", "Расчет инвестиций", "NEW CNR", "Расчет инвестиций (2)"]
        sheet = None
        sheet_name = None
        for name in target_sheet_names:
            if name in workbook.sheetnames:
                sheet = workbook[name]
                sheet_name = name
                print(f"Найден лист с данными планирования: {name}")
                break
        if not sheet:
            # Если не нашли по имени, пробуем первый лист
            sheet_name = workbook.sheetnames[0]
            sheet = workbook[sheet_name]
            print(f"Используется первый доступный лист: {sheet_name}")

        # Маркеры начала секций (уникальные для каждой)
        start_marker_sales = "Блок ПЛАНИРОВАНИЕ продаж"
        # Маркеры начала *других* секций, которые обозначают конец текущей
        other_section_markers = [
            "Распределение инвестиций контракта, учитываемые в ЦМ, %", # Начало инвестиций
            "ПЛАНИРОВАНИЕ ИНВЕСТИЦИЙ", # Заголовок в строке 85 или 225, 241, 284 (если он уникален как начало)
            # Добавьте другие маркеры, если известны
        ]

        start_row_idx = None
        end_row_idx = None

        print("Поиск маркера начала блока 'Блок ПЛАНИРОВАНИЕ продаж'...")
        for row_idx, row in enumerate(sheet.iter_rows(values_only=False), 1):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if start_marker_sales in cell.value and start_row_idx is None:
                        start_row_idx = row_idx
                        print(f"Найден маркер начала '{start_marker_sales}' в строке {start_row_idx}")
                        break # Нашли начало, выходим из внутреннего цикла
            if start_row_idx is not None:
                break # Нашли начало, выходим из внешнего цикла

        if start_row_idx is None:
            print(f"Маркер начала '{start_marker_sales}' не найден")
            return None, 0

        print("Поиск маркера конца (начала следующей секции)...")
        # Теперь ищем маркер начала *другой* секции *после* найденного начала
        for row_idx in range(start_row_idx + 1, sheet.max_row + 1):
            first_cell = sheet.cell(row=row_idx, column=1) # Проверяем только первый столбец
            if first_cell.value and isinstance(first_cell.value, str):
                cell_val_upper = first_cell.value.strip().upper()
                # Проверяем, не является ли первая ячейка строки началом другой секции
                for marker in other_section_markers:
                    if marker.strip().upper() in cell_val_upper:
                        end_row_idx = row_idx - 1 # Конец - строка перед началом другой секции
                        print(f"Найден маркер начала другой секции '{first_cell.value}' в строке {row_idx}, извлекаем до {end_row_idx}")
                        break
                if end_row_idx is not None:
                    break # Нашли конец, выходим из цикла

        if end_row_idx is None:
            print("Маркер конца не найден, извлекаем данные до конца листа")
            end_row_idx = sheet.max_row

        # Проверяем, что диапазон корректен
        if end_row_idx < start_row_idx:
            print(f"Некорректный диапазон: start={start_row_idx}, end={end_row_idx}")
            return None, 0

        # Найдем строку с "ПЛАНИРОВАНИЕ ПРОДАЖ" в первом столбце как потенциальный заголовок данных
        data_start_row = start_row_idx
        for r in range(start_row_idx, min(end_row_idx + 1, sheet.max_row + 1)):
            first_cell = sheet.cell(row=r, column=1)
            if first_cell.value and isinstance(first_cell.value, str) and first_cell.value.strip().upper() == "ПЛАНИРОВАНИЕ ПРОДАЖ":
                data_start_row = r + 1 # Данные начинаются со следующей строки
                print(f"Найден заголовок 'ПЛАНИРОВАНИЕ ПРОДАЖ' в строке {r}, данные начинаются с {data_start_row}")
                break
        # Если заголовок "ПЛАНИРОВАНИЕ ПРОДАЖ" не найден, начинаем с строки после маркера начала
        if data_start_row == start_row_idx:
             data_start_row = start_row_idx + 1 # Пропускаем строку с "Блок ПЛАНИРОВАНИЕ продаж"

        data_end_row = end_row_idx

        print(f"Извлечение данных планирования продаж с {data_start_row} по {data_end_row} строку...")

        # Создаем новый рабочий лист
        excel_wb = openpyxl.Workbook()
        excel_ws = excel_wb.active
        excel_ws.title = "Планирование продаж"

        # Добавляем заголовок
        excel_ws.merge_cells('A1:Z1')
        header_cell = excel_ws['A1']
        header_cell.value = "ПЛАНИРОВАНИЕ ПРОДАЖ"
        header_cell.font = Font(bold=True, size=16)
        header_cell.alignment = Alignment(horizontal='center')
        excel_ws.merge_cells('A2:Z2')
        excel_ws['A2'] = f"Дата извлечения: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Лист: {sheet_name} | Диапазон: с {data_start_row} по {data_end_row}"
        excel_ws['A2'].font = Font(italic=True)
        excel_ws['A2'].alignment = Alignment(horizontal='center')

        # Добавляем отступ
        current_excel_row = 4

        # Копируем данные в новый файл
        rows_copied = 0
        for r in range(data_start_row, data_end_row + 1):
            has_data = False
            for c in range(1, sheet.max_column + 1):
                source_cell = sheet.cell(row=r, column=c)
                target_cell = excel_ws.cell(row=current_excel_row, column=c)
                # Копируем значение
                if source_cell.value is not None:
                    target_cell.value = source_cell.value
                    has_data = True
                # Копируем стиль (безопасный способ, как в других функциях)
                if source_cell.has_style:
                    # Копируем шрифт
                    if source_cell.font:
                        try:
                            target_cell.font = Font(
                                name=source_cell.font.name,
                                size=source_cell.font.size,
                                bold=source_cell.font.bold,
                                italic=source_cell.font.italic,
                                underline=source_cell.font.underline,
                                color=source_cell.font.color
                            )
                        except:
                            target_cell.font = Font()
                    # Копируем границы
                    if source_cell.border:
                        try:
                            target_cell.border = Border(
                                left=source_cell.border.left,
                                right=source_cell.border.right,
                                top=source_cell.border.top,
                                bottom=source_cell.border.bottom
                            )
                        except:
                            target_cell.border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin')
                            )
                    # Копируем заливку
                    if source_cell.fill:
                        try:
                            if source_cell.fill.fill_type == "solid" and source_cell.fill.fgColor:
                                try:
                                    rgb = source_cell.fill.fgColor.rgb
                                    target_cell.fill = PatternFill(
                                        fill_type="solid",
                                        fgColor=rgb
                                    )
                                except:
                                    target_cell.fill = PatternFill(
                                        fill_type=source_cell.fill.fill_type,
                                        fgColor="FFFFFF"
                                    )
                            else:
                                target_cell.fill = source_cell.fill
                        except:
                            target_cell.fill = PatternFill(fill_type=None)
                    # Копируем выравнивание
                    if source_cell.alignment:
                        try:
                            target_cell.alignment = Alignment(
                                horizontal=source_cell.alignment.horizontal,
                                vertical=source_cell.alignment.vertical,
                                wrap_text=source_cell.alignment.wrap_text,
                                indent=source_cell.alignment.indent
                            )
                        except:
                            target_cell.alignment = Alignment(horizontal='left', vertical='center')
                    # Копируем числовой формат
                    if hasattr(source_cell, 'number_format') and source_cell.number_format:
                        try:
                            target_cell.number_format = source_cell.number_format
                        except:
                            target_cell.number_format = 'General'
            # Переходим к следующей строке только если были данные
            if has_data:
                current_excel_row += 1
                rows_copied += 1
            # Иначе пропускаем пустую строку (но не увеличиваем current_excel_row)

        print(f"Скопировано {rows_copied} строк с данными")

        # Если не было данных, возвращаем None
        if rows_copied == 0:
            print("Не найдено данных для извлечения")
            return None, 0

        # Добавляем рамку вокруг данных
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        for r in range(4, current_excel_row):
            for c in range(1, sheet.max_column + 1):
                cell = excel_ws.cell(row=r, column=c)
                # Если у ячейки нет границ, добавляем рамку
                if not cell.border or (not cell.border.left.style and not cell.border.right.style and
                                      not cell.border.top.style and not cell.border.bottom.style):
                    cell.border = thin_border

        # Автоподбор ширины столбцов
        for col in range(1, sheet.max_column + 1):
            max_length = 0
            column = get_column_letter(col)
            for row in range(4, current_excel_row):
                cell = excel_ws.cell(row=row, column=col)
                if cell.value:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            adjusted_width = (max_length + 2)
            if adjusted_width > 50:
                adjusted_width = 50
            excel_ws.column_dimensions[column].width = adjusted_width

        # Добавляем подвал
        excel_ws.merge_cells(f'A{current_excel_row}:Z{current_excel_row}')
        excel_ws[f'A{current_excel_row}'] = f"Отчет сгенерирован автоматически {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        excel_ws[f'A{current_excel_row}'].font = Font(italic=True, size=10)
        excel_ws[f'A{current_excel_row}'].alignment = Alignment(horizontal='right')

        return excel_wb, rows_copied

    except Exception as e:
        print(f"Произошла ошибка при извлечении данных планирования продаж: {e}")
        traceback.print_exc()
        return None, 0


def extract_investment_planning_data(file_path, output_dir):
    """
    Ищет и извлекает данные между маркерами:
    "Распределение инвестиций контракта, учитываемые в ЦМ, %" (уникальный маркер начала)
    и
    началом следующей секции (например, "Блок ПЛАНИРОВАНИЕ продаж") или концом листа.
    """
    try:
        print("--- Поиск данных планирования инвестиций ---")
        # Загружаем рабочую книгу
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        # Ищем лист с нужными данными
        target_sheet_names = ["NEW CNR 1", "Расчет инвестиций", "NEW CNR", "Расчет инвестиций (2)"]
        sheet = None
        sheet_name = None
        for name in target_sheet_names:
            if name in workbook.sheetnames:
                sheet = workbook[name]
                sheet_name = name
                print(f"Найден лист с данными инвестиций: {name}")
                break
        if not sheet:
            # Если не нашли по имени, пробуем первый лист
            sheet_name = workbook.sheetnames[0]
            sheet = workbook[sheet_name]
            print(f"Используется первый доступный лист: {sheet_name}")

        # Маркеры начала секций (уникальные для каждой)
        start_marker_investment = "Распределение инвестиций контракта, учитываемые в ЦМ, %"
        # Маркеры начала *других* секций, которые обозначают конец текущей
        other_section_markers = [
            "Блок ПЛАНИРОВАНИЕ продаж", # Начало продаж
            # Добавьте другие маркеры, если известны
        ]

        start_row_idx = None
        end_row_idx = None

        print("Поиск маркера начала блока 'Распределение инвестиций контракта, учитываемые в ЦМ, %'...")
        for row_idx, row in enumerate(sheet.iter_rows(values_only=False), 1):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if start_marker_investment in cell.value and start_row_idx is None:
                        start_row_idx = row_idx
                        print(f"Найден маркер начала '{start_marker_investment}' в строке {start_row_idx}")
                        break # Нашли начало, выходим из внутреннего цикла
            if start_row_idx is not None:
                break # Нашли начало, выходим из внешнего цикла

        if start_row_idx is None:
            print(f"Маркер начала '{start_marker_investment}' не найден")
            # Попробуем найти "ПЛАНИРОВАНИЕ ИНВЕСТИЦИЙ" в начале строки как альтернативу, но с осторожностью
            print("Поиск альтернативного маркера 'ПЛАНИРОВАНИЕ ИНВЕСТИЦИЙ' в начале строки...")
            for row_idx in range(1, sheet.max_row + 1):
                first_cell = sheet.cell(row=row_idx, column=1)
                if first_cell.value and isinstance(first_cell.value, str) and first_cell.value.strip().upper() == "ПЛАНИРОВАНИЕ ИНВЕСТИЦИЙ":
                    # Проверим, является ли это строка 85 (уникальное начало) или одна из строк 225, 241, 284 (не начало)
                    # Проверим контекст: строка 85 обычно после "Условия контракта", а строки 225, 241, 284 - внутри таблицы продаж
                    # Проверим, есть ли перед этой строкой "Условия контракта" или что-то другое
                    # Проверим строки перед ней
                    context_found = False
                    for check_row in range(max(1, row_idx - 10), row_idx): # Проверим 10 строк перед
                        check_cell = sheet.cell(row=check_row, column=1)
                        if check_cell.value and isinstance(check_cell.value, str):
                            if "УСЛОВИЯ" in check_cell.value.upper():
                                # Нашли "Условия" перед этим заголовком - возможно это начало
                                # Проверим, не встречается ли "Блок ПЛАНИРОВАНИЕ продаж" между "Условия" и этим заголовком
                                sales_block_found = False
                                for check_between_row in range(check_row + 1, row_idx):
                                    check_between_cell = sheet.cell(row=check_between_row, column=1)
                                    if check_between_cell.value and isinstance(check_between_cell.value, str) and "БЛОК ПЛАНИРОВАНИЕ ПРОДАЖ" in check_between_cell.value:
                                        sales_block_found = True
                                        break
                                if not sales_block_found:
                                    start_row_idx = row_idx
                                    print(f"Найден альтернативный маркер начала 'ПЛАНИРОВАНИЕ ИНВЕСТИЦИЙ' в строке {start_row_idx} (после 'Условия', до 'Блок Планирование')")
                                    context_found = True
                                    break
                    if context_found:
                        break # Нашли подходящий альтернативный маркер

            if start_row_idx is None:
                print("Альтернативный маркер также не найден")
                return None, 0

        print("Поиск маркера конца (начала следующей секции)...")
        # Теперь ищем маркер начала *другой* секции *после* найденного начала
        for row_idx in range(start_row_idx + 1, sheet.max_row + 1):
            first_cell = sheet.cell(row=row_idx, column=1) # Проверяем только первый столбец
            if first_cell.value and isinstance(first_cell.value, str):
                cell_val_upper = first_cell.value.strip().upper()
                # Проверяем, не является ли первая ячейка строки началом другой секции
                for marker in other_section_markers:
                    if marker.strip().upper() in cell_val_upper:
                        end_row_idx = row_idx - 1 # Конец - строка перед началом другой секции
                        print(f"Найден маркер начала другой секции '{first_cell.value}' в строке {row_idx}, извлекаем до {end_row_idx}")
                        break
                if end_row_idx is not None:
                    break # Нашли конец, выходим из цикла

        if end_row_idx is None:
            print("Маркер конца не найден, извлекаем данные до конца листа")
            end_row_idx = sheet.max_row

        # Проверяем, что диапазон корректен
        if end_row_idx < start_row_idx:
            print(f"Некорректный диапазон: start={start_row_idx}, end={end_row_idx}")
            return None, 0

        # Найдем строку с "ПЛАНИРОВАНИЕ ИНВЕСТИЦИЙ" в первом столбце как потенциальный заголовок данных
        data_start_row = start_row_idx
        for r in range(start_row_idx, min(end_row_idx + 1, sheet.max_row + 1)):
            first_cell = sheet.cell(row=r, column=1)
            if first_cell.value and isinstance(first_cell.value, str) and first_cell.value.strip().upper() == "ПЛАНИРОВАНИЕ ИНВЕСТИЦИЙ":
                data_start_row = r + 1 # Данные начинаются со следующей строки
                print(f"Найден заголовок 'ПЛАНИРОВАНИЕ ИНВЕСТИЦИЙ' в строке {r}, данные начинаются с {data_start_row}")
                break
        # Если заголовок "ПЛАНИРОВАНИЕ ИНВЕСТИЦИЙ" не найден, начинаем с строки после маркера начала
        if data_start_row == start_row_idx:
             data_start_row = start_row_idx + 1 # Пропускаем строку с "Распределение инвестиций контракта..."

        data_end_row = end_row_idx

        print(f"Извлечение данных планирования инвестиций с {data_start_row} по {data_end_row} строку...")

        # Создаем новый рабочий лист
        excel_wb = openpyxl.Workbook()
        excel_ws = excel_wb.active
        excel_ws.title = "Планирование инвестиций"

        # Добавляем заголовок
        excel_ws.merge_cells('A1:Z1')
        header_cell = excel_ws['A1']
        header_cell.value = "ПЛАНИРОВАНИЕ ИНВЕСТИЦИЙ"
        header_cell.font = Font(bold=True, size=16)
        header_cell.alignment = Alignment(horizontal='center')
        excel_ws.merge_cells('A2:Z2')
        excel_ws['A2'] = f"Дата извлечения: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Лист: {sheet_name} | Диапазон: с {data_start_row} по {data_end_row}"
        excel_ws['A2'].font = Font(italic=True)
        excel_ws['A2'].alignment = Alignment(horizontal='center')

        # Добавляем отступ
        current_excel_row = 4

        # Копируем данные в новый файл
        rows_copied = 0
        for r in range(data_start_row, data_end_row + 1):
            has_data = False
            for c in range(1, sheet.max_column + 1):
                source_cell = sheet.cell(row=r, column=c)
                target_cell = excel_ws.cell(row=current_excel_row, column=c)
                # Копируем значение
                if source_cell.value is not None:
                    target_cell.value = source_cell.value
                    has_data = True
                # Копируем стиль (используем безопасную функцию, как в других функциях)
                if source_cell.has_style:
                    # Копируем шрифт
                    if source_cell.font:
                        try:
                            target_cell.font = Font(
                                name=source_cell.font.name,
                                size=source_cell.font.size,
                                bold=source_cell.font.bold,
                                italic=source_cell.font.italic,
                                underline=source_cell.font.underline,
                                color=source_cell.font.color
                            )
                        except:
                            target_cell.font = Font()
                    # Копируем границы
                    if source_cell.border:
                        try:
                            target_cell.border = Border(
                                left=source_cell.border.left,
                                right=source_cell.border.right,
                                top=source_cell.border.top,
                                bottom=source_cell.border.bottom
                            )
                        except:
                            target_cell.border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin')
                            )
                    # Копируем заливку
                    if source_cell.fill:
                        try:
                            if source_cell.fill.fill_type == "solid" and source_cell.fill.fgColor:
                                try:
                                    rgb = source_cell.fill.fgColor.rgb
                                    target_cell.fill = PatternFill(
                                        fill_type="solid",
                                        fgColor=rgb
                                    )
                                except:
                                    target_cell.fill = PatternFill(
                                        fill_type=source_cell.fill.fill_type,
                                        fgColor="FFFFFF"
                                    )
                            else:
                                target_cell.fill = source_cell.fill
                        except:
                            target_cell.fill = PatternFill(fill_type=None)
                    # Копируем выравнивание
                    if source_cell.alignment:
                        try:
                            target_cell.alignment = Alignment(
                                horizontal=source_cell.alignment.horizontal,
                                vertical=source_cell.alignment.vertical,
                                wrap_text=source_cell.alignment.wrap_text,
                                indent=source_cell.alignment.indent
                            )
                        except:
                            target_cell.alignment = Alignment(horizontal='left', vertical='center')
                    # Копируем числовой формат
                    if hasattr(source_cell, 'number_format') and source_cell.number_format:
                        try:
                            target_cell.number_format = source_cell.number_format
                        except:
                            target_cell.number_format = 'General'
            # Переходим к следующей строке только если были данные
            if has_data:
                current_excel_row += 1
                rows_copied += 1

        print(f"Скопировано {rows_copied} строк с данными")

        # Если не было данных, возвращаем None
        if rows_copied == 0:
            print("Не найдено данных для извлечения")
            return None, 0

        # Добавляем рамку вокруг данных (используем простой и надежный способ, как в других функциях)
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        for r in range(4, current_excel_row):
            for c in range(1, sheet.max_column + 1):
                cell = excel_ws.cell(row=r, column=c)
                # Если у ячейки нет границ, добавляем рамку
                if not cell.border or (not cell.border.left.style and not cell.border.right.style and
                                      not cell.border.top.style and not cell.border.bottom.style):
                    cell.border = thin_border

        # Автоподбор ширины столбцов
        for col in range(1, sheet.max_column + 1):
            max_length = 0
            column = get_column_letter(col)
            for row in range(4, current_excel_row):
                cell = excel_ws.cell(row=row, column=col)
                if cell.value:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            adjusted_width = (max_length + 2)
            if adjusted_width > 50:
                adjusted_width = 50
            excel_ws.column_dimensions[column].width = adjusted_width

        # Добавляем подвал
        excel_ws.merge_cells(f'A{current_excel_row}:Z{current_excel_row}')
        excel_ws[f'A{current_excel_row}'] = f"Отчет сгенерирован автоматически {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        excel_ws[f'A{current_excel_row}'].font = Font(italic=True, size=10)
        excel_ws[f'A{current_excel_row}'].alignment = Alignment(horizontal='right')

        return excel_wb, rows_copied

    except Exception as e:
        print(f"Произошла ошибка при извлечении данных планирования инвестиций: {e}")
        traceback.print_exc()
        return None, 0


def safe_copy_style(source_cell, target_cell):
    """
    Безопасно копирует стили из одной ячейки в другую
    """
    try:
        # Копируем шрифт
        if source_cell.font:
            try:
                # Получаем атрибуты шрифта
                font_attrs = {}
                if hasattr(source_cell.font, 'name') and source_cell.font.name:
                    font_attrs['name'] = source_cell.font.name
                if hasattr(source_cell.font, 'size') and source_cell.font.size:
                    font_attrs['size'] = source_cell.font.size
                if hasattr(source_cell.font, 'bold') and source_cell.font.bold is not None:
                    font_attrs['bold'] = source_cell.font.bold
                if hasattr(source_cell.font, 'italic') and source_cell.font.italic is not None:
                    font_attrs['italic'] = source_cell.font.italic
                if hasattr(source_cell.font, 'underline') and source_cell.font.underline is not None:
                    font_attrs['underline'] = source_cell.font.underline
                if hasattr(source_cell.font, 'color') and source_cell.font.color:
                    font_attrs['color'] = source_cell.font.color
                # Создаем новый объект шрифта
                target_cell.font = Font(**font_attrs)
            except Exception as e:
                # Если не удалось скопировать шрифт, используем стандартный
                target_cell.font = Font()
        # Копируем границы
        if source_cell.border:
            try:
                # Получаем атрибуты границ
                border_attrs = {}
                for side_name in ['left', 'right', 'top', 'bottom']:
                    side = getattr(source_cell.border, side_name, None)
                    if side and hasattr(side, 'style') and side.style:
                        border_attrs[side_name] = Side(style=side.style)
                # Создаем новый объект границ
                target_cell.border = Border(**border_attrs)
            except Exception as e:
                # Если не удалось скопировать границы, используем тонкие границы
                target_cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        # Копируем заливку
        if source_cell.fill:
            try:
                # Получаем атрибуты заливки
                fill_attrs = {}
                if hasattr(source_cell.fill, 'fill_type') and source_cell.fill.fill_type:
                    fill_attrs['fill_type'] = source_cell.fill.fill_type
                if hasattr(source_cell.fill, 'fgColor') and source_cell.fill.fgColor:
                    try:
                        # Пытаемся получить RGB цвет
                        if hasattr(source_cell.fill.fgColor, 'rgb') and source_cell.fill.fgColor.rgb:
                            fill_attrs['fgColor'] = source_cell.fill.fgColor.rgb
                        elif hasattr(source_cell.fill.fgColor, 'tint') and source_cell.fill.fgColor.tint:
                            fill_attrs['fgColor'] = source_cell.fill.fgColor.tint
                    except:
                        pass
                # Создаем новый объект заливки
                if fill_attrs:
                    target_cell.fill = PatternFill(**fill_attrs)
                else:
                    target_cell.fill = PatternFill(fill_type=None)
            except Exception as e:
                # Если не удалось скопировать заливку, оставляем без заливки
                target_cell.fill = PatternFill(fill_type=None)
        # Копируем выравнивание
        if source_cell.alignment:
            try:
                # Получаем атрибуты выравнивания
                alignment_attrs = {}
                if hasattr(source_cell.alignment, 'horizontal') and source_cell.alignment.horizontal:
                    alignment_attrs['horizontal'] = source_cell.alignment.horizontal
                if hasattr(source_cell.alignment, 'vertical') and source_cell.alignment.vertical:
                    alignment_attrs['vertical'] = source_cell.alignment.vertical
                if hasattr(source_cell.alignment, 'wrap_text') and source_cell.alignment.wrap_text is not None:
                    alignment_attrs['wrap_text'] = source_cell.alignment.wrap_text
                if hasattr(source_cell.alignment, 'indent') and source_cell.alignment.indent is not None:
                    alignment_attrs['indent'] = source_cell.alignment.indent
                # Создаем новый объект выравнивания
                target_cell.alignment = Alignment(**alignment_attrs)
            except Exception as e:
                # Если не удалось скопировать выравнивание, используем стандартное
                target_cell.alignment = Alignment(horizontal='left', vertical='center')
        # Копируем числовой формат
        if hasattr(source_cell, 'number_format') and source_cell.number_format:
            try:
                target_cell.number_format = source_cell.number_format
            except:
                target_cell.number_format = 'General'
    except Exception as e:
        # Если произошла ошибка при копировании стилей, просто игнорируем
        pass

def merge_tables(gfd_wb, contract_wb, planning_wb, investment_wb, source_file_path, output_path):
    """
    Объединяет четыре таблицы в один Excel файл и копирует лист "SAP-код"
    """
    try:
        # Создаем новый файл для объединенных данных
        merged_wb = openpyxl.Workbook()
        # Удаляем лист по умолчанию
        default_sheet = merged_wb.active
        merged_wb.remove(default_sheet)
        # Копируем лист GFD таблицы
        if gfd_wb:
            gfd_ws = list(gfd_wb.worksheets)[0]
            gfd_sheet = merged_wb.create_sheet(title="GFD Запрос")
            # Копируем данные с безопасным копированием стилей
            for row in gfd_ws.iter_rows(values_only=False):
                for cell in row:
                    new_cell = gfd_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                    safe_copy_style(cell, new_cell)
        # Копируем лист таблицы условий контракта
        if contract_wb:
            contract_ws = list(contract_wb.worksheets)[0]
            contract_sheet = merged_wb.create_sheet(title="Условия контракта")
            # Копируем данные с безопасным копированием стилей
            for row in contract_ws.iter_rows(values_only=False):
                for cell in row:
                    new_cell = contract_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                    safe_copy_style(cell, new_cell)
        # Копируем лист планирования продаж
        if planning_wb:
            planning_ws = list(planning_wb.worksheets)[0]
            planning_sheet = merged_wb.create_sheet(title="Планирование продаж")
            # Копируем данные с безопасным копированием стилей
            for row in planning_ws.iter_rows(values_only=False):
                for cell in row:
                    new_cell = planning_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                    safe_copy_style(cell, new_cell)
        # Копируем лист планирования инвестиций
        if investment_wb:
            investment_ws = list(investment_wb.worksheets)[0]
            investment_sheet = merged_wb.create_sheet(title="Планирование инвестиций")
            # Копируем данные с безопасным копированием стилей
            for row in investment_ws.iter_rows(values_only=False):
                for cell in row:
                    new_cell = investment_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                    safe_copy_style(cell, new_cell)
        # --- ДОБАВЛЕННЫЙ БЛОК: Копируем лист "SAP-код" из исходного файла ---
        try:
            source_wb = openpyxl.load_workbook(source_file_path, data_only=False)
            if "SAP-код" in source_wb.sheetnames:
                sap_sheet = source_wb["SAP-код"]
                new_sap_sheet = merged_wb.create_sheet(title="SAP-код")
                # Копируем все данные и стили
                for row in sap_sheet.iter_rows(values_only=False):
                    for cell in row:
                        new_cell = new_sap_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                        safe_copy_style(cell, new_cell)
                # Копируем размеры столбцов
                for col in sap_sheet.column_dimensions:
                    if col in new_sap_sheet.column_dimensions:
                        new_sap_sheet.column_dimensions[col].width = sap_sheet.column_dimensions[col].width
                # Копируем размеры строк
                for row_num in sap_sheet.row_dimensions:
                    if row_num in new_sap_sheet.row_dimensions:
                        new_sap_sheet.row_dimensions[row_num].height = sap_sheet.row_dimensions[row_num].height
                print("Лист 'SAP-код' успешно скопирован в итоговый файл.")
            else:
                print("Лист 'SAP-код' не найден в исходном файле.")
        except Exception as e:
            print(f"Ошибка при копировании листа 'SAP-код': {e}")
            traceback.print_exc()
        # --- КОНЕЦ ДОБАВЛЕННОГО БЛОКА ---
        # Добавляем сводный лист
        summary_sheet = merged_wb.create_sheet(title="Сводка", index=0)
        # Заголовок
        summary_sheet.merge_cells('A1:D1')
        summary_sheet['A1'] = "ОБЪЕДИНЕННЫЙ ОТЧЕТ ПО КОНТРАКТАМ"
        summary_sheet['A1'].font = Font(bold=True, size=16)
        summary_sheet['A1'].alignment = Alignment(horizontal='center')
        # Информация
        summary_sheet['A3'] = "Таблица 1:"
        summary_sheet['A3'].font = Font(bold=True)
        summary_sheet['B3'] = "Запрос на заключение контракта по напиткам GFD"
        summary_sheet['A4'] = "Таблица 2:"
        summary_sheet['A4'].font = Font(bold=True)
        summary_sheet['B4'] = "Условия контракта"
        summary_sheet['A5'] = "Таблица 3:"
        summary_sheet['A5'].font = Font(bold=True)
        summary_sheet['B5'] = "Планирование продаж"
        summary_sheet['A6'] = "Таблица 4:"
        summary_sheet['A6'].font = Font(bold=True)
        summary_sheet['B6'] = "Планирование инвестиций"
        summary_sheet['A7'] = "Лист 5:"
        summary_sheet['A7'].font = Font(bold=True)
        summary_sheet['B7'] = "SAP-код"
        summary_sheet['A9'] = "Для просмотра данных перейдите на соответствующие листы:"
        summary_sheet['A10'] = "- 'GFD Запрос' - содержит данные о запросе на заключение контракта"
        summary_sheet['A11'] = "- 'Условия контракта' - содержит условия контракта"
        summary_sheet['A12'] = "- 'Планирование продаж' - содержит данные планирования продаж"
        summary_sheet['A13'] = "- 'Планирование инвестиций' - содержит данные по инвестициям"
        summary_sheet['A14'] = "- 'SAP-код' - содержит коды SAP"
        # Форматирование
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        for row in range(3, 15):
            for col in range(1, 5):
                summary_sheet.cell(row=row, column=col).border = thin_border
        # Автоподбор ширины
        summary_sheet.column_dimensions['A'].width = 15
        summary_sheet.column_dimensions['B'].width = 50
        summary_sheet.column_dimensions['C'].width = 15
        summary_sheet.column_dimensions['D'].width = 15
        # Сохраняем результат
        merged_wb.save(output_path)
        print(f"Объединенный отчет успешно сохранен: {output_path}")
        return True
    except Exception as e:
        print(f"Произошла ошибка при объединении таблиц: {e}")
        traceback.print_exc()
        return False


def extract_and_merge_tables(file_path, output_dir):
    """
    Извлекает и объединяет все таблицы, включая данные планирования инвестиций.
    Итоговый файл сохраняется под оригинальным именем исходного файла.
    """
    # Извлекаем таблицу GFD запроса
    print("--- Извлечение таблицы GFD запроса ---")
    gfd_wb, gfd_rows = extract_gfd_request_table(file_path, output_dir)
    if not gfd_wb:
        print("Не удалось извлечь таблицу GFD запроса")
    else:
        print(f"Извлечено {gfd_rows} строк данных из таблицы GFD запроса")

    # Извлекаем таблицу условий контракта
    print("\n--- Извлечение таблицы условий контракта ---")
    contract_wb, contract_rows = extract_contract_conditions_table(file_path, output_dir)
    if not contract_wb:
        print("Не удалось извлечь таблицу условий контракта")
    else:
        print(f"Извлечено {contract_rows} строк данных из таблицы условий контракта")

    # Извлекаем данные планирования продаж
    print("\n--- Извлечение данных планирования продаж ---")
    planning_wb, planning_rows = extract_planning_sales_data(file_path, output_dir)
    if not planning_wb:
        print("Не удалось извлечь данные планирования продаж")
    else:
        print(f"Извлечено {planning_rows} строк данных планирования продаж")

    # Извлекаем данные планирования инвестиций
    print("\n--- Извлечение данных планирования инвестиций ---")
    investment_wb, investment_rows = extract_investment_planning_data(file_path, output_dir)
    if not investment_wb:
        print("Не удалось извлечь данные планирования инвестиций")
    else:
        print(f"Извлечено {investment_rows} строк данных планирования инвестиций")

    # Объединяем таблицы
    # Формируем имя выходного файла: используем ИСХОДНОЕ имя файла
    base_name = os.path.basename(file_path)
    output_filename = base_name  # Сохраняем под оригинальным именем
    output_path = os.path.join(output_dir, output_filename)
    
    return merge_tables(gfd_wb, contract_wb, planning_wb, investment_wb, file_path, output_path)

# Основной код
if __name__ == "__main__":
    # Пути из задания
    file_dir = r"C:\Users\metelkov\Desktop\эцп тест\file_new\file"
    output_dir = r"C:\Users\metelkov\Desktop\эцп тест\file_new\map"
    # Проверяем существование директорий
    if not os.path.exists(file_dir):
        print(f"Директория {file_dir} не существует!")
        exit(1)
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"Создана директория: {output_dir}")
    # Получаем список всех файлов в директории
    all_files = os.listdir(file_dir)
    # Фильтруем, оставляя только Excel файлы
    excel_files = [f for f in all_files if f.endswith('.xlsx') or f.endswith('.xls')]
    total_files = len(excel_files)
    print(f"Найдено {total_files} Excel файлов для обработки")
    if not excel_files:
        print("В указанной директории не найдено Excel файлов")
        exit(1)
    # Инициализируем списки для итогового отчета
    success_files = []
    failed_files = []
    # Создаем прогресс-бар
    with tqdm(total=total_files, desc="Обработка файлов", unit="файл") as pbar:
        start_time = time.time()
        # Обрабатываем каждый файл
        for idx, file_name in enumerate(excel_files, 1):
            file_path = os.path.join(file_dir, file_name)
            print(f"\n{'='*60}")
            print(f"📄 [{idx}/{total_files}] Обработка файла: {file_name}")
            print(f"{'='*60}")
            # Проверяем существование файла
            if not os.path.exists(file_path):
                print(f"❌ Файл {file_path} не существует!")
                failed_files.append(file_name)
                pbar.update(1)
                continue
            # Извлекаем и объединяем таблицы
            print("⏳ Начинаем процесс извлечения и объединения таблиц...")
            try:
                success = extract_and_merge_tables(file_path, output_dir)
                if success:
                    success_files.append(file_name)
                    print(f"✅ Файл '{file_name}' успешно обработан.")
                else:
                    failed_files.append(file_name)
                    print(f"❌ Файл '{file_name}' обработан с ошибками.")
            except Exception as e:
                failed_files.append(file_name)
                print(f"❌ Критическая ошибка при обработке файла '{file_name}': {e}")
                traceback.print_exc()
            # Обновляем прогресс-бар
            pbar.update(1)
            # Расчет примерного оставшегося времени
            elapsed_time = time.time() - start_time
            avg_time_per_file = elapsed_time / idx
            remaining_files = total_files - idx
            estimated_remaining_time = avg_time_per_file * remaining_files
            pbar.set_postfix({
                'Успешно': len(success_files),
                'Ошибки': len(failed_files),
                'Осталось': f"{estimated_remaining_time:.1f}с"
            })

    # Итоговая статистика
    print("\n" + "="*70)
    print("📊 ЗАВЕРШЕНИЕ ОБРАБОТКИ ВСЕХ ФАЙЛОВ")
    print("="*70)
    print(f"Всего файлов обработано: {total_files}")
    print(f"✅ Успешно: {len(success_files)}")
    print(f"❌ С ошибками: {len(failed_files)}")
    if success_files:
        print("\n📋 Список успешно обработанных файлов:")
        for fname in success_files:
            print(f"  - {fname}")
    if failed_files:
        print("\n📋 Список файлов, которые не удалось обработать:")
        for fname in failed_files:
            print(f"  - {fname}")
    if len(success_files) > 0:
        print(f"\n🎉 Процесс завершен успешно. Итоговые файлы сохранены в: {output_dir}")
    else:
        print(f"\n⚠️  Не удалось успешно обработать ни один файл.")